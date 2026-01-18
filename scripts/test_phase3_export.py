"""
Test script to verify Phase 3 Excel export changes.
Run this to generate a test export and verify the changes are working.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src', 'backend'))

from app import create_app
from models import db
from sqlalchemy import text
from app.services.phase3_planning_service import Phase3PlanningService
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

def test_export(organization_id=48):
    """Generate a test export to verify changes."""

    app = create_app()

    with app.app_context():
        service = Phase3PlanningService(db.session)
        output = service.get_phase3_output(organization_id)
        config = output.get('config', {})
        summary = output.get('summary', {})
        modules = output.get('modules', [])
        timeline = output.get('timeline', {})

        # Get org name
        org_result = db.session.execute(
            text("SELECT organization_name FROM organization WHERE id = :org_id"),
            {'org_id': organization_id}
        ).fetchone()
        org_name = org_result.organization_name if org_result else f'Organization_{organization_id}'

        # Get scaling info from correct location
        training_modules_data = output.get('training_modules', {})
        scaling_info = training_modules_data.get('scaling_info', {})

        selected_view = config.get('selected_view', 'competency_level')
        is_role_clustered = selected_view == 'role_clustered'

        # Style definitions
        HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        HEADER_FONT = Font(bold=True, color='FFFFFF')
        TITLE_FONT = Font(size=16, bold=True)
        SUBTITLE_FONT = Font(size=11, bold=True)
        LABEL_FONT = Font(bold=True)
        CLUSTER_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        CLUSTER_FONT = Font(bold=True, size=11)
        THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        LEVEL_NAMES = {1: 'Knowing', 2: 'Understanding', 4: 'Applying'}

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Phase 3 - Macro Planning'

        row = 1

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'SE-QPT Phase 3: Macro Planning Export (TEST)'
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='center')
        row = 3

        # Summary section
        ws[f'A{row}'] = 'Organization:'
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'B{row}'] = org_name
        row += 1

        ws[f'A{row}'] = 'Training View:'
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'B{row}'] = 'Role-Clustered' if is_role_clustered else 'Competency-Level'
        row += 1

        ws[f'A{row}'] = 'Total Training Modules:'
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'B{row}'] = summary.get('total_modules', 0)
        row += 2

        # ===== SCALED PARTICIPANT ESTIMATION NOTE =====
        factor = scaling_info.get('scaling_factor', 1)
        print(f"[CHECK] Scaling factor: {factor}")
        print(f"[CHECK] Should show note: {factor > 1}")

        if scaling_info and factor and factor > 1:
            NOTE_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            NOTE_BORDER = Border(
                left=Side(style='thin', color='E6A23C'),
                right=Side(style='thin', color='E6A23C'),
                top=Side(style='thin', color='E6A23C'),
                bottom=Side(style='thin', color='E6A23C')
            )

            ws[f'A{row}'] = 'Note: Scaled Participant Estimation'
            ws[f'A{row}'].font = Font(bold=True, color='E6A23C')
            row += 1

            assessed = scaling_info.get('actual_assessed_users', 0)
            target = scaling_info.get('target_group_size', 0)

            note_lines = [
                f"Only {assessed} out of {target} employees in the target group completed the competency assessment.",
                f"Participant counts are scaled by a factor of {factor:.1f}x to estimate organization-wide training needs.",
                "Actual participation may vary based on individual competency gaps and role assignments.",
                "These estimates should be used for planning purposes and may require adjustment."
            ]

            for note_line in note_lines:
                cell = ws[f'A{row}']
                cell.value = f"  - {note_line}"
                cell.fill = NOTE_FILL
                cell.border = NOTE_BORDER
                ws.merge_cells(f'A{row}:F{row}')
                row += 1

            row += 1
            print("[SUCCESS] Scaling note added to export")
        else:
            print("[SKIP] No scaling note needed (factor <= 1)")

        # ===== TRAINING MODULES TABLE =====
        ws[f'A{row}'] = 'Training Modules'
        ws[f'A{row}'].font = SUBTITLE_FONT
        row += 1

        def format_module_name(competency_name, pmt_type):
            if not pmt_type or pmt_type.lower() in ['combined', '-', 'null', 'none', '']:
                return competency_name
            return f"{competency_name} - {pmt_type.capitalize()}"

        def get_cluster_name(cluster_id):
            if not cluster_id:
                return None
            result = db.session.execute(
                text("SELECT training_program_name FROM training_program_cluster WHERE id = :id"),
                {'id': cluster_id}
            ).fetchone()
            return result.training_program_name if result else None

        # Table headers
        if is_role_clustered:
            headers = ['Training Program', 'Module Type', 'Training Module', 'Level', 'Learning Format', 'Est. Participants']
            col_widths = [22, 18, 40, 15, 18, 16]
        else:
            headers = ['Training Module', 'Level', 'Learning Format', 'Est. Participants']
            col_widths = [50, 15, 20, 18]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

        row += 1

        if is_role_clustered:
            COMMON_BASE_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            ROLE_SPECIFIC_FILL = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')

            # Group by cluster
            cluster_groups = {}
            for module in modules:
                cluster_display = module.get('cluster_name')
                if not cluster_display or cluster_display == 'Uncategorized':
                    cluster_id = module.get('cluster_id') or module.get('training_program_cluster_id')
                    if cluster_id:
                        cluster_display = get_cluster_name(cluster_id)
                if not cluster_display:
                    cluster_display = 'Uncategorized'

                if cluster_display not in cluster_groups:
                    cluster_groups[cluster_display] = []
                cluster_groups[cluster_display].append(module)

            # Order clusters
            cluster_order = ['SE for Engineers', 'SE for Managers', 'SE for Interfacing Partners']
            sorted_clusters = []
            for cluster in cluster_order:
                if cluster in cluster_groups:
                    sorted_clusters.append((cluster, cluster_groups[cluster]))
            for cluster, mods in cluster_groups.items():
                if cluster not in cluster_order:
                    sorted_clusters.append((cluster, mods))

            # Write modules
            for cluster_display, cluster_modules in sorted_clusters:
                cluster_start_row = row

                # For Engineers, use subcluster
                if 'Engineer' in cluster_display:
                    common_base = [m for m in cluster_modules if m.get('subcluster') == 'common']
                    role_specific = [m for m in cluster_modules if m.get('subcluster') != 'common']

                    print(f"[CHECK] {cluster_display}: Common Base={len(common_base)}, Role-Specific={len(role_specific)}")

                    # Common Base modules
                    for module in common_base:
                        level_num = module.get('target_level', 0)
                        level_name = LEVEL_NAMES.get(level_num, f'Level {level_num}')
                        module_name = format_module_name(module.get('competency_name', ''), module.get('pmt_type', ''))
                        format_id = module.get('selected_format_id')
                        format_name = 'Not Selected'
                        if format_id:
                            fmt_result = db.session.execute(
                                text("SELECT format_name FROM learning_format WHERE id = :id"),
                                {'id': format_id}
                            ).fetchone()
                            if fmt_result:
                                format_name = fmt_result.format_name

                        values = [cluster_display, 'Common Base', module_name, level_name, format_name,
                                  module.get('estimated_participants', 0)]
                        for col_idx, value in enumerate(values, 1):
                            cell = ws.cell(row=row, column=col_idx, value=value)
                            cell.border = THIN_BORDER
                            if col_idx == 1:
                                cell.fill = CLUSTER_FILL
                                cell.font = CLUSTER_FONT
                            if col_idx == 2:
                                cell.fill = COMMON_BASE_FILL
                            if col_idx in [4, 6]:
                                cell.alignment = Alignment(horizontal='center')
                        row += 1

                    # Role-Specific modules
                    for module in role_specific:
                        level_num = module.get('target_level', 0)
                        level_name = LEVEL_NAMES.get(level_num, f'Level {level_num}')
                        module_name = format_module_name(module.get('competency_name', ''), module.get('pmt_type', ''))

                        # Add role info
                        pathway_roles = module.get('pathway_roles') or module.get('roles_needing_training', [])
                        if pathway_roles and isinstance(pathway_roles, list):
                            role_info = ', '.join(pathway_roles[:2])
                            if len(pathway_roles) > 2:
                                role_info += f' +{len(pathway_roles)-2} more'
                            module_name = f"{module_name} ({role_info})"

                        format_id = module.get('selected_format_id')
                        format_name = 'Not Selected'
                        if format_id:
                            fmt_result = db.session.execute(
                                text("SELECT format_name FROM learning_format WHERE id = :id"),
                                {'id': format_id}
                            ).fetchone()
                            if fmt_result:
                                format_name = fmt_result.format_name

                        values = [cluster_display, 'Role-Specific', module_name, level_name, format_name,
                                  module.get('estimated_participants', 0)]
                        for col_idx, value in enumerate(values, 1):
                            cell = ws.cell(row=row, column=col_idx, value=value)
                            cell.border = THIN_BORDER
                            if col_idx == 1:
                                cell.fill = CLUSTER_FILL
                                cell.font = CLUSTER_FONT
                            if col_idx == 2:
                                cell.fill = ROLE_SPECIFIC_FILL
                            if col_idx in [4, 6]:
                                cell.alignment = Alignment(horizontal='center')
                        row += 1
                else:
                    # Non-Engineers
                    for module in cluster_modules:
                        level_num = module.get('target_level', 0)
                        level_name = LEVEL_NAMES.get(level_num, f'Level {level_num}')
                        module_name = format_module_name(module.get('competency_name', ''), module.get('pmt_type', ''))
                        format_id = module.get('selected_format_id')
                        format_name = 'Not Selected'
                        if format_id:
                            fmt_result = db.session.execute(
                                text("SELECT format_name FROM learning_format WHERE id = :id"),
                                {'id': format_id}
                            ).fetchone()
                            if fmt_result:
                                format_name = fmt_result.format_name

                        values = [cluster_display, '-', module_name, level_name, format_name,
                                  module.get('estimated_participants', 0)]
                        for col_idx, value in enumerate(values, 1):
                            cell = ws.cell(row=row, column=col_idx, value=value)
                            cell.border = THIN_BORDER
                            if col_idx == 1:
                                cell.fill = CLUSTER_FILL
                                cell.font = CLUSTER_FONT
                            if col_idx in [4, 6]:
                                cell.alignment = Alignment(horizontal='center')
                        row += 1

                # Merge cluster cells
                if row - cluster_start_row > 1:
                    ws.merge_cells(f'A{cluster_start_row}:A{row - 1}')
                    ws[f'A{cluster_start_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Save file
        output_path = os.path.join(os.path.dirname(__file__), '..', 'Phase3_Test_Export.xlsx')
        wb.save(output_path)
        print(f"\n[SUCCESS] Test export saved to: {os.path.abspath(output_path)}")

        return output_path


if __name__ == '__main__':
    org_id = 48
    if len(sys.argv) > 1:
        org_id = int(sys.argv[1])

    print(f"Testing Phase 3 export for organization {org_id}...")
    print()
    test_export(org_id)
