"""
Phase 3: Macro Planning Routes Blueprint

API endpoints for Phase 3 "Macro Planning":
- Task 1: Training Structure Selection
- Task 2: Learning Format Selection with 3-Factor Suitability
- Task 3: LLM-Generated Timeline Planning

Based on Phase3_Macro_Planning_Specification_v3.2.md
"""

from flask import Blueprint, request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
import traceback

from models import db

# Create blueprint
phase3_planning_bp = Blueprint('phase3_planning', __name__)


# ==============================================================================
# HELPER: Get Service Instance
# ==============================================================================

def _get_service():
    """Get Phase 3 Planning Service instance"""
    from app.services.phase3_planning_service import Phase3PlanningService
    return Phase3PlanningService(db.session)


# ==============================================================================
# TASK 1: Training Structure Selection
# ==============================================================================

@phase3_planning_bp.route('/phase3/config/<int:organization_id>', methods=['GET'])
@jwt_required()
def get_phase3_config(organization_id):
    """
    Get Phase 3 configuration for an organization.

    Returns config including selected view, progress, and scaling info.
    """
    try:
        service = _get_service()

        config = service.get_phase3_config(organization_id)
        available_views = service.get_available_views(organization_id)

        return jsonify({
            'success': True,
            'config': config,
            'available_views': available_views
        })

    except Exception as e:
        current_app.logger.error(f"Error getting Phase 3 config: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/training-structure/<int:organization_id>', methods=['GET'])
@jwt_required()
def get_training_structure(organization_id):
    """
    Get training structure options for Task 1.

    Returns available views based on organization maturity and current selection.
    """
    try:
        service = _get_service()

        config = service.get_phase3_config(organization_id)
        available_views = service.get_available_views(organization_id)

        return jsonify({
            'success': True,
            'selected_view': config.get('selected_view'),
            'task_completed': config.get('task1_completed', False),
            **available_views
        })

    except Exception as e:
        current_app.logger.error(f"Error getting training structure: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/training-structure/<int:organization_id>', methods=['POST'])
@jwt_required()
def set_training_structure(organization_id):
    """
    Set training structure view for Task 1.

    Request Body:
        {"selected_view": "competency_level" | "role_clustered"}
    """
    try:
        data = request.get_json()
        selected_view = data.get('selected_view')

        if not selected_view:
            return jsonify({
                'success': False,
                'error': 'selected_view is required'
            }), 400

        service = _get_service()
        result = service.set_training_structure(organization_id, selected_view)

        return jsonify(result)

    except ValueError as ve:
        return jsonify({
            'success': False,
            'error': str(ve)
        }), 400
    except Exception as e:
        current_app.logger.error(f"Error setting training structure: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# TASK 2: Learning Format Selection
# ==============================================================================

@phase3_planning_bp.route('/phase3/learning-formats', methods=['GET'])
@jwt_required()
def get_learning_formats():
    """
    Get all 10 learning formats with their properties.

    Returns complete format definitions including participant ranges,
    max achievable levels, and effort ratings.
    """
    try:
        service = _get_service()
        formats = service.get_learning_formats()

        return jsonify({
            'success': True,
            'formats': formats,
            'total': len(formats)
        })

    except Exception as e:
        current_app.logger.error(f"Error getting learning formats: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/training-modules/<int:organization_id>', methods=['GET'])
@jwt_required()
def get_training_modules(organization_id):
    """
    Get training modules from Phase 2 Learning Objectives.

    Query Params:
        view_type: 'competency_level' (default) or 'role_clustered'

    Returns modules with gap data, participant estimates, and existing selections.
    For role_clustered view, returns separate modules per training cluster.
    """
    try:
        view_type = request.args.get('view_type', 'competency_level')

        service = _get_service()
        result = service.get_training_modules(organization_id, view_type=view_type)

        if result.get('error'):
            return jsonify({
                'success': False,
                'error': result['error']
            }), 400

        return jsonify({
            'success': True,
            **result
        })

    except Exception as e:
        current_app.logger.error(f"Error getting training modules: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/evaluate-format', methods=['POST'])
@jwt_required()
def evaluate_format():
    """
    Evaluate a learning format's suitability for a training module.

    Request Body:
        {
            "organization_id": 28,
            "competency_id": 14,
            "target_level": 2,
            "format_id": 1,
            "participant_count": 45
        }

    Returns 3-factor suitability feedback (green/yellow/red for each).
    """
    try:
        data = request.get_json()

        required = ['organization_id', 'competency_id', 'target_level', 'format_id', 'participant_count']
        missing = [f for f in required if f not in data]
        if missing:
            return jsonify({
                'success': False,
                'error': f"Missing required fields: {missing}"
            }), 400

        service = _get_service()
        result = service.evaluate_format_suitability(
            organization_id=data['organization_id'],
            competency_id=data['competency_id'],
            target_level=data['target_level'],
            format_id=data['format_id'],
            participant_count=data['participant_count']
        )

        return jsonify({
            'success': True,
            **result
        })

    except ValueError as ve:
        return jsonify({
            'success': False,
            'error': str(ve)
        }), 400
    except Exception as e:
        current_app.logger.error(f"Error evaluating format: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/select-format', methods=['POST'])
@jwt_required()
def select_format():
    """
    Save format selection for a training module.

    Request Body:
        {
            "organization_id": 28,
            "competency_id": 14,
            "target_level": 2,
            "pmt_type": "combined" | "method" | "tool",
            "format_id": 1,
            "estimated_participants": 45,
            "confirmed": true,
            "cluster_id": null | 1  // Optional: for role_clustered view
        }
    """
    try:
        data = request.get_json()

        required = ['organization_id', 'competency_id', 'target_level', 'pmt_type', 'format_id']
        missing = [f for f in required if f not in data]
        if missing:
            return jsonify({
                'success': False,
                'error': f"Missing required fields: {missing}"
            }), 400

        service = _get_service()

        # First evaluate suitability
        suitability = service.evaluate_format_suitability(
            organization_id=data['organization_id'],
            competency_id=data['competency_id'],
            target_level=data['target_level'],
            format_id=data['format_id'],
            participant_count=data.get('estimated_participants', 0)
        )

        # Save selection (with optional cluster_id for role_clustered view)
        result = service.save_format_selection(
            organization_id=data['organization_id'],
            competency_id=data['competency_id'],
            target_level=data['target_level'],
            pmt_type=data['pmt_type'],
            format_id=data['format_id'],
            suitability=suitability,
            estimated_participants=data.get('estimated_participants', 0),
            confirmed=data.get('confirmed', False),
            cluster_id=data.get('cluster_id')
        )

        return jsonify({
            'success': True,
            'suitability': suitability
        })

    except Exception as e:
        current_app.logger.error(f"Error saving format selection: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/complete-task2/<int:organization_id>', methods=['POST'])
@jwt_required()
def complete_task2(organization_id):
    """
    Mark Task 2 (Learning Format Selection) as completed.
    Called when user clicks 'Continue to Task 3' after configuring all modules.
    """
    try:
        service = _get_service()
        result = service.mark_task2_completed(organization_id)
        return jsonify({
            'success': True,
            'message': 'Task 2 marked as completed'
        })
    except Exception as e:
        current_app.logger.error(f"Error completing Task 2: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# TASK 3: Timeline Planning
# ==============================================================================

@phase3_planning_bp.route('/phase3/generate-timeline', methods=['POST'])
@jwt_required()
def generate_timeline():
    """
    Generate timeline milestones using LLM.

    Request Body:
        {"organization_id": 28}

    Returns 5 milestones with estimated dates.
    """
    try:
        data = request.get_json()
        organization_id = data.get('organization_id')

        if not organization_id:
            return jsonify({
                'success': False,
                'error': 'organization_id is required'
            }), 400

        service = _get_service()
        result = service.generate_timeline(organization_id)

        if not result.get('success'):
            return jsonify(result), 500

        return jsonify(result)

    except Exception as e:
        current_app.logger.error(f"Error generating timeline: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/timeline/<int:organization_id>', methods=['GET'])
@jwt_required()
def get_timeline(organization_id):
    """
    Get stored timeline milestones for an organization.
    """
    try:
        service = _get_service()
        result = service.get_timeline(organization_id)

        return jsonify({
            'success': True,
            **result
        })

    except Exception as e:
        current_app.logger.error(f"Error getting timeline: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# Phase 3 Output/Summary
# ==============================================================================

@phase3_planning_bp.route('/phase3/output/<int:organization_id>', methods=['GET'])
@jwt_required()
def get_phase3_output(organization_id):
    """
    Get complete Phase 3 output summary.

    Returns all Phase 3 data including config, modules, timeline, and summary stats.
    """
    try:
        service = _get_service()
        result = service.get_phase3_output(organization_id)

        return jsonify({
            'success': True,
            **result
        })

    except Exception as e:
        current_app.logger.error(f"Error getting Phase 3 output: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# Training Program Clusters (for Role-Clustered View)
# ==============================================================================

@phase3_planning_bp.route('/phase3/training-clusters', methods=['GET'])
@jwt_required()
def get_training_clusters():
    """
    Get all 6 Training Program Clusters.

    Used for Role-Clustered training view.
    """
    try:
        result = db.session.execute(
            db.text("""
                SELECT id, cluster_key, cluster_name, training_program_name,
                       description, typical_org_roles
                FROM training_program_cluster
                ORDER BY display_order
            """)
        )

        clusters = []
        for row in result:
            clusters.append({
                'id': row.id,
                'cluster_key': row.cluster_key,
                'cluster_name': row.cluster_name,
                'training_program_name': row.training_program_name,
                'description': row.description,
                'typical_org_roles': row.typical_org_roles
            })

        return jsonify({
            'success': True,
            'clusters': clusters,
            'total': len(clusters)
        })

    except Exception as e:
        current_app.logger.error(f"Error getting training clusters: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@phase3_planning_bp.route('/phase3/training-clusters/<int:organization_id>/distribution', methods=['GET'])
@jwt_required()
def get_training_cluster_distribution(organization_id):
    """
    Get distribution of organization roles across Training Program Clusters.

    Used for Role-Clustered training view.
    Note: Uses organization_roles table (not organization_role_mappings)
    """
    try:
        result = db.session.execute(
            db.text("""
                SELECT tpc.id, tpc.cluster_name, tpc.training_program_name,
                       COUNT(orr.id) as role_count,
                       ARRAY_AGG(orr.role_name) as role_titles
                FROM training_program_cluster tpc
                LEFT JOIN organization_roles orr
                    ON orr.training_program_cluster_id = tpc.id
                    AND orr.organization_id = :org_id
                GROUP BY tpc.id, tpc.cluster_name, tpc.training_program_name
                ORDER BY tpc.display_order
            """),
            {'org_id': organization_id}
        )

        distribution = []
        for row in result:
            distribution.append({
                'cluster_id': row.id,
                'cluster_name': row.cluster_name,
                'training_program_name': row.training_program_name,
                'role_count': row.role_count,
                'role_titles': [r for r in row.role_titles if r] if row.role_titles else []
            })

        return jsonify({
            'success': True,
            'distribution': distribution
        })

    except Exception as e:
        current_app.logger.error(f"Error getting cluster distribution: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==============================================================================
# Export to Excel
# ==============================================================================

@phase3_planning_bp.route('/phase3/export/<int:organization_id>', methods=['GET'])
@jwt_required()
def export_phase3_excel(organization_id):
    """
    Export Phase 3 Macro Planning data to Excel.

    Only available after all 3 tasks are completed.
    Single sheet with Summary header + Training Modules table.
    For Role-Clustered view: groups modules by Training Program Cluster.
    """
    from flask import send_file
    from datetime import datetime
    import io

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({
            'success': False,
            'error': 'Excel export not available',
            'message': 'openpyxl library not installed'
        }), 500

    try:
        service = _get_service()

        # Get Phase 3 output data
        output = service.get_phase3_output(organization_id)
        config = output.get('config', {})
        summary = output.get('summary', {})
        modules = output.get('modules', [])
        timeline = output.get('timeline', {})

        # Check if all tasks are completed
        completion = summary.get('completion', {})
        if not (completion.get('task1') and completion.get('task2') and completion.get('task3')):
            return jsonify({
                'success': False,
                'error': 'Export only available after completing all Phase 3 tasks'
            }), 400

        # Get organization name and maturity score
        org_result = db.session.execute(
            db.text("SELECT organization_name, maturity_score FROM organization WHERE id = :org_id"),
            {'org_id': organization_id}
        ).fetchone()
        org_name = org_result.organization_name if org_result else f'Organization_{organization_id}'
        maturity_score = org_result.maturity_score if org_result and org_result.maturity_score else 0
        if maturity_score >= 80: maturity_level = 5
        elif maturity_score >= 60: maturity_level = 4
        elif maturity_score >= 40: maturity_level = 3
        elif maturity_score >= 20: maturity_level = 2
        else: maturity_level = 1

        # Get strategy descriptions for selected strategies
        from app.services.strategy_selection_engine import SE_TRAINING_STRATEGIES
        strategy_desc_map = {s['name']: s.get('description', '') for s in SE_TRAINING_STRATEGIES}
        strategy_detail_map = {s['name']: s for s in SE_TRAINING_STRATEGIES}

        # Get full strategy objects (with priorities)
        all_strategy_objs = service._get_all_strategies(organization_id)

        # Determine view type
        selected_view = config.get('selected_view', 'competency_level')
        is_role_clustered = selected_view == 'role_clustered'

        # Style definitions
        HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        HEADER_FONT = Font(bold=True, color='FFFFFF')
        TITLE_FONT = Font(size=16, bold=True)
        SUBTITLE_FONT = Font(size=11, bold=True)
        LABEL_FONT = Font(bold=True)
        CLUSTER_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        CLUSTER_FONT = Font(bold=True, size=11)
        THIN_BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Competency topic header row
        COMP_HEADER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        COMP_HEADER_FONT = Font(bold=True, italic=True, size=10)

        # Package header row fills (role-clustered only)
        PKG_ENGINEERS_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        PKG_MANAGERS_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        PKG_PARTNERS_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        PKG_HEADER_FONT = Font(bold=True, size=11)

        # Subcluster header row (Engineers only)
        SUBCLUSTER_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        SUBCLUSTER_FONT = Font(bold=True, italic=True, size=10, color='606060')

        # Level name mapping
        LEVEL_NAMES = {1: 'Knowing', 2: 'Understanding', 4: 'Applying', 6: 'Mastering'}
        LEVEL_DURATION_HOURS = {1: 1, 2: 2, 4: 4, 6: 8}

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Phase 3 - Macro Planning'

        row = 1

        # ===== TITLE =====
        ws.merge_cells('A1:H1')
        ws['A1'] = 'SE-QPT Phase 3: Macro Planning Export'
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='center')
        row = 3

        # Style for description/annotation rows
        DESC_FONT = Font(italic=True, size=9, color='606266')
        PHASE_FONT = Font(italic=True, size=9, color='909399')

        # Helper: write an overview item with optional description and phase origin
        def write_overview_item(r, label, value, phase_origin=None, description=None, merge_value=True):
            ws[f'A{r}'] = label
            ws[f'A{r}'].font = LABEL_FONT
            ws[f'B{r}'] = value
            if merge_value:
                ws.merge_cells(f'B{r}:F{r}')
            if phase_origin:
                ws[f'G{r}'] = phase_origin
                ws[f'G{r}'].font = PHASE_FONT
            r += 1
            if description:
                ws[f'B{r}'] = description
                ws[f'B{r}'].font = DESC_FONT
                ws.merge_cells(f'B{r}:H{r}')
                r += 1
            return r

        # ===== OVERVIEW SECTION =====
        ws[f'A{row}'] = 'Overview'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        # Organization
        row = write_overview_item(row, 'Organization:', org_name,
            description='The organization for which this SE qualification plan was created.')

        # SE Maturity Level
        row = write_overview_item(row, 'SE Maturity Level:', f'Level {maturity_level}/5 (Score: {maturity_score:.1f}/100)',
            phase_origin='(Phase 1, Task 3)',
            description='Organization SE maturity assessed via the Maturity Assessment questionnaire.')

        # Target Group Size
        target_size = summary.get('target_group_size', 0)
        target_range = summary.get('target_group_range', '')
        target_display = f"{target_size} participants"
        if target_range and target_range != 'Unknown':
            target_display += f" ({target_range})"
        row = write_overview_item(row, 'Target Group Size:', target_display,
            phase_origin='(Phase 1, Task 1)',
            description='Number of employees in the target group to be qualified in SE competencies.')

        # Selected Strategies (with descriptions)
        all_strategies = summary.get('all_strategies', [summary.get('strategy_name', 'Not Selected')])
        strategy_names_str = ', '.join(all_strategies) if all_strategies else 'Not Selected'
        # Add primary/secondary labels
        if all_strategy_objs:
            labeled = []
            for s in all_strategy_objs:
                lbl = f"{s['name']} (Primary)" if s.get('is_primary') else s['name']
                labeled.append(lbl)
            strategy_names_str = ', '.join(labeled)
        row = write_overview_item(row, 'Selected Strategies:', strategy_names_str,
            phase_origin='(Phase 1, Task 2)',
            description='Qualification strategies selected to guide the training program design.')

        # Add individual strategy descriptions
        for s_name in all_strategies:
            desc = strategy_desc_map.get(s_name, '')
            detail = strategy_detail_map.get(s_name, {})
            if desc:
                # Strategy name as sub-label
                ws[f'B{row}'] = f"{s_name}:"
                ws[f'B{row}'].font = Font(bold=True, size=9)
                row += 1
                ws[f'B{row}'] = desc
                ws[f'B{row}'].font = DESC_FONT
                ws.merge_cells(f'B{row}:H{row}')
                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                row += 1
                # Add key details in one line
                details_parts = []
                if detail.get('targetAudience'):
                    details_parts.append(f"Target: {detail['targetAudience']}")
                if detail.get('qualificationLevel'):
                    details_parts.append(f"Level: {detail['qualificationLevel']}")
                if detail.get('duration'):
                    details_parts.append(f"Duration: {detail['duration']}")
                if details_parts:
                    ws[f'B{row}'] = ' | '.join(details_parts)
                    ws[f'B{row}'].font = Font(italic=True, size=9, color='909399')
                    ws.merge_cells(f'B{row}:H{row}')
                    row += 1

        # Training View
        view_label = 'Role-Clustered' if is_role_clustered else 'Competency-Level'
        if is_role_clustered:
            view_desc = 'Modules grouped into 3 training packages (SE for Engineers, SE for Managers, SE for Interfacing Partners) based on role cluster assignments.'
        else:
            view_desc = 'Modules grouped by the 16 SE competency areas, each containing sub-modules at identified competency levels.'
        row = write_overview_item(row, 'Training View:', view_label,
            phase_origin='(Phase 3, Task 1)',
            description=view_desc)

        # Total Training Modules
        total_modules = summary.get('total_modules', 0)
        row = write_overview_item(row, 'Total Training Modules:', total_modules,
            phase_origin='(Phase 3, Task 2)',
            description='Training modules identified from competency gap analysis (Phase 2) and learning format assignment.')

        # Total Estimated Duration
        total_duration = sum(LEVEL_DURATION_HOURS.get(m.get('target_level', 0), 2) for m in modules)
        duration_display = f"{total_duration} hours"
        if total_duration >= 8:
            duration_display += f" (~{total_duration / 8:.1f} training days)"
        row = write_overview_item(row, 'Total Est. Duration:', duration_display,
            phase_origin='(Derived)',
            description='Cumulative estimated training duration across all modules (L1=1h, L2=2h, L4=4h, L6=8h per module).')

        # Format Distribution
        format_dist = summary.get('format_distribution', {})
        if format_dist:
            format_str = ', '.join([f"{k}: {v}" for k, v in format_dist.items()])
            row = write_overview_item(row, 'Format Distribution:', format_str,
                phase_origin='(Phase 3, Task 2)',
                description='Distribution of learning formats selected for training delivery (e.g., Seminar, WBT, Blended Learning).')

        # Export Date
        ws[f'A{row}'] = 'Export Date:'
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        row += 2

        # Set column widths for overview area
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14

        # ===== SCALED PARTICIPANT ESTIMATION NOTE =====
        # Get scaling info from training_modules
        training_modules_data = output.get('training_modules', {})
        scaling_info = training_modules_data.get('scaling_info', {})
        if scaling_info and scaling_info.get('scaling_factor') and scaling_info.get('scaling_factor', 1) > 1:
            NOTE_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            NOTE_BORDER = Border(
                left=Side(style='thin', color='E6A23C'),
                right=Side(style='thin', color='E6A23C'),
                top=Side(style='thin', color='E6A23C'),
                bottom=Side(style='thin', color='E6A23C')
            )

            ws[f'A{row}'] = 'Note: Scaled Participant Estimation'
            ws[f'A{row}'].font = Font(bold=True, color='E6A23C')
            row += 1

            assessed = scaling_info.get('actual_assessed_users', 0)
            target = scaling_info.get('target_group_size', 0)
            factor = scaling_info.get('scaling_factor', 1)

            note_lines = [
                f"Only {assessed} out of {target} employees in the target group completed the competency assessment.",
                f"Participant counts are scaled by a factor of {factor:.1f}x to estimate organization-wide training needs.",
                "Actual participation may vary based on individual competency gaps and role assignments.",
                "These estimates should be used for planning purposes and may require adjustment."
            ]

            for note_line in note_lines:
                cell = ws[f'A{row}']
                cell.value = f"  - {note_line}"
                cell.fill = NOTE_FILL
                cell.border = NOTE_BORDER
                ws.merge_cells(f'A{row}:H{row}')
                row += 1

            row += 1

        # ===== TRAINING MODULES TABLE =====
        ws[f'A{row}'] = 'Training Modules'
        ws[f'A{row}'].font = SUBTITLE_FONT
        row += 1

        # Helper to format module name (Competency + PMT Type)
        def format_module_name(competency_name, pmt_type):
            """Format module name: 'Competency - PMT' or just 'Competency' if combined/null"""
            if not pmt_type or pmt_type.lower() in ['combined', '-', 'null', 'none', '']:
                return competency_name
            return f"{competency_name} - {pmt_type.capitalize()}"

        # Helper to get cluster name from cluster_id
        def get_cluster_name(cluster_id):
            """Look up cluster name from training_program_cluster table"""
            if not cluster_id:
                return None
            result = db.session.execute(
                db.text("SELECT training_program_name FROM training_program_cluster WHERE id = :id"),
                {'id': cluster_id}
            ).fetchone()
            return result.training_program_name if result else None

        # Table headers
        if is_role_clustered:
            headers = ['Training Program', 'Module Type', 'Training Module', 'Level', 'Learning Format', 'Roles', 'Est. Participants', 'Est. Duration (h)']
            col_widths = [20, 14, 32, 14, 28, 45, 14, 14]
        else:
            headers = ['Training Module', 'Level', 'Learning Format', 'Roles', 'Est. Participants', 'Est. Duration (h)']
            col_widths = [35, 14, 30, 50, 14, 14]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
            _cl = get_column_letter(col_idx)
            ws.column_dimensions[_cl].width = max(col_widths[col_idx - 1], ws.column_dimensions[_cl].width or 0)

        header_row = row
        row += 1

        # Helper: write a merged section header row
        def write_section_header(row_num, text, fill, font, num_cols):
            merge_range = f'A{row_num}:{get_column_letter(num_cols)}{row_num}'
            ws.merge_cells(merge_range)
            cell = ws[f'A{row_num}']
            cell.value = text
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical='center')
            return row_num + 1

        # Helper: format roles list
        def format_roles(module):
            roles = module.get('pathway_roles') or module.get('roles_needing_training', [])
            if roles and isinstance(roles, list):
                return ', '.join(roles)
            return ''

        # Helper: resolve format name for a module
        def resolve_format_name(module):
            fname = module.get('format_name', '')
            if not fname and module.get('selected_format_id'):
                fmt_result = db.session.execute(
                    db.text("SELECT format_name FROM learning_format WHERE id = :id"),
                    {'id': module['selected_format_id']}
                ).fetchone()
                fname = fmt_result.format_name if fmt_result else 'Not Selected'
            elif not fname:
                fname = 'Not Selected'
            return fname

        # Helper: write a single module data row (role-clustered)
        def write_rc_module_row(row_num, module, idx, cluster_display, module_type):
            level_num = module.get('target_level', 0)
            level_name = LEVEL_NAMES.get(level_num, f'Level {level_num}')
            mod_name = format_module_name(
                module.get('competency_name', ''),
                module.get('pmt_type', '')
            )
            fname = resolve_format_name(module)
            roles_str = format_roles(module)
            duration = LEVEL_DURATION_HOURS.get(level_num, 2)
            values = [idx, cluster_display, module_type, mod_name, level_name, fname,
                      roles_str, module.get('estimated_participants', 0), duration]
            for ci, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=ci, value=value)
                cell.border = THIN_BORDER
                if ci in [1, 5, 8, 9]:
                    cell.alignment = Alignment(horizontal='center')
                if ci == 7:
                    cell.alignment = Alignment(wrap_text=True)
            return row_num + 1

        # Helper: write competency-grouped modules within a section
        from itertools import groupby

        def write_competency_groups(row_num, section_modules, cluster_display, module_type, idx_counter, num_cols):
            sorted_mods = sorted(section_modules, key=lambda m: m.get('competency_id', 0))
            for _, grp_iter in groupby(sorted_mods, key=lambda m: m.get('competency_id', 0)):
                grp = list(grp_iter)
                comp_name = grp[0].get('competency_name', 'Unknown')
                count = len(grp)
                label = f"    {comp_name} ({count} sub-module{'s' if count != 1 else ''})"
                row_num = write_section_header(row_num, label, COMP_HEADER_FILL, COMP_HEADER_FONT, num_cols)
                for mod in grp:
                    idx_counter[0] += 1
                    row_num = write_rc_module_row(row_num, mod, idx_counter[0], cluster_display, module_type)
            return row_num

        # Group modules by cluster for Role-Clustered view
        if is_role_clustered:
            # Update headers to include # column
            headers = ['#', 'Training Program', 'Module Type', 'Training Module', 'Level', 'Learning Format', 'Roles', 'Est. Participants', 'Est. Duration (h)']
            col_widths = [4, 20, 14, 32, 14, 28, 45, 14, 14]
            # Re-write header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = THIN_BORDER
                _cl = get_column_letter(col_idx)
                ws.column_dimensions[_cl].width = max(col_widths[col_idx - 1], ws.column_dimensions[_cl].width or 0)

            num_cols = len(headers)

            # Package fill mapping
            pkg_fills = {
                'SE for Engineers': PKG_ENGINEERS_FILL,
                'SE for Managers': PKG_MANAGERS_FILL,
                'SE for Interfacing Partners': PKG_PARTNERS_FILL,
            }

            # Sort modules by cluster - use cluster_id to look up proper name
            cluster_groups = {}
            for module in modules:
                cluster_display = module.get('cluster_name')
                if not cluster_display or cluster_display == 'Uncategorized':
                    cluster_id = module.get('cluster_id') or module.get('training_program_cluster_id')
                    if cluster_id:
                        cluster_display = get_cluster_name(cluster_id)
                if not cluster_display:
                    cluster_display = 'Uncategorized'
                if cluster_display not in cluster_groups:
                    cluster_groups[cluster_display] = []
                cluster_groups[cluster_display].append(module)

            cluster_order = ['SE for Engineers', 'SE for Managers', 'SE for Interfacing Partners']
            sorted_clusters = []
            for cluster in cluster_order:
                if cluster in cluster_groups:
                    sorted_clusters.append((cluster, cluster_groups[cluster]))
            for cluster, mods in cluster_groups.items():
                if cluster not in cluster_order:
                    sorted_clusters.append((cluster, mods))

            idx_counter = [0]  # mutable counter for sequential numbering

            for cluster_display, cluster_modules in sorted_clusters:
                # Write package header row
                pkg_fill = pkg_fills.get(cluster_display, CLUSTER_FILL)
                row = write_section_header(row, cluster_display, pkg_fill, PKG_HEADER_FONT, num_cols)

                if 'Engineer' in cluster_display:
                    common_base = [m for m in cluster_modules if m.get('subcluster') == 'common']
                    role_specific = [m for m in cluster_modules if m.get('subcluster') != 'common']

                    if common_base:
                        cb_label = f"  Common Base ({len(common_base)} module{'s' if len(common_base) != 1 else ''})"
                        row = write_section_header(row, cb_label, SUBCLUSTER_FILL, SUBCLUSTER_FONT, num_cols)
                        row = write_competency_groups(row, common_base, cluster_display, 'Common Base', idx_counter, num_cols)

                    if role_specific:
                        rs_label = f"  Role-Specific Pathways ({len(role_specific)} module{'s' if len(role_specific) != 1 else ''})"
                        row = write_section_header(row, rs_label, SUBCLUSTER_FILL, SUBCLUSTER_FONT, num_cols)
                        row = write_competency_groups(row, role_specific, cluster_display, 'Role-Specific', idx_counter, num_cols)
                else:
                    # Managers / Interfacing Partners - no subcluster split
                    row = write_competency_groups(row, cluster_modules, cluster_display, '-', idx_counter, num_cols)

        else:
            # Standard competency-level view - group by competency
            num_cols = len(headers)  # headers already set above (5 cols)
            sorted_mods = sorted(modules, key=lambda m: m.get('competency_id', 0))
            idx = 0

            for _, grp_iter in groupby(sorted_mods, key=lambda m: m.get('competency_id', 0)):
                grp = list(grp_iter)
                comp_name = grp[0].get('competency_name', 'Unknown')
                count = len(grp)
                label = f"{comp_name} ({count} sub-module{'s' if count != 1 else ''})"
                row = write_section_header(row, label, COMP_HEADER_FILL, COMP_HEADER_FONT, num_cols)

                for module in grp:
                    idx += 1
                    level_num = module.get('target_level', 0)
                    level_name = LEVEL_NAMES.get(level_num, f'Level {level_num}')
                    module_name = format_module_name(
                        module.get('competency_name', ''),
                        module.get('pmt_type', '')
                    )
                    fname = resolve_format_name(module)
                    roles = module.get('roles_needing_training', [])
                    roles_str = ', '.join(roles) if roles and isinstance(roles, list) else ''
                    duration = LEVEL_DURATION_HOURS.get(level_num, 2)

                    values = [
                        module_name,
                        level_name,
                        fname,
                        roles_str,
                        module.get('estimated_participants', 0),
                        duration
                    ]

                    for col_idx, value in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        cell.border = THIN_BORDER
                        if col_idx in [2, 5, 6]:
                            cell.alignment = Alignment(horizontal='center')
                        if col_idx == 4:
                            cell.alignment = Alignment(wrap_text=True)

                    row += 1

        row += 2

        # ===== TIMELINE SECTION =====
        milestones = timeline.get('milestones', [])
        if milestones:
            ws[f'A{row}'] = 'Implementation Timeline'
            ws[f'A{row}'].font = SUBTITLE_FONT
            row += 1

            # Timeline headers
            timeline_headers = ['#', 'Milestone', 'Date', 'Quarter', 'Description']
            timeline_widths = [5, 30, 15, 12, 70]

            for col_idx, header in enumerate(timeline_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = THIN_BORDER
                _cl = get_column_letter(col_idx)
                _cur = ws.column_dimensions[_cl].width or 0
                if timeline_widths[col_idx - 1] > _cur:
                    ws.column_dimensions[_cl].width = timeline_widths[col_idx - 1]

            row += 1

            for idx, milestone in enumerate(milestones, 1):
                values = [
                    idx,
                    milestone.get('milestone_name', milestone.get('name', '')),
                    milestone.get('estimated_date', ''),
                    milestone.get('quarter', ''),
                    milestone.get('milestone_description', milestone.get('description', ''))
                ]

                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = THIN_BORDER
                    if col_idx in [1, 3, 4]:
                        cell.alignment = Alignment(horizontal='center')
                    if col_idx == 5:
                        cell.alignment = Alignment(wrap_text=True)

                row += 1

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        safe_org_name = ''.join(c for c in org_name if c.isalnum() or c in ' -_')[:30]
        timestamp = datetime.now().strftime('%Y%m%d')
        filename = f"Phase3_MacroPlanning_{safe_org_name}_{timestamp}.xlsx"

        current_app.logger.info(f"[Phase3 Export] Excel export created for org {organization_id}")

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"Error exporting Phase 3: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
