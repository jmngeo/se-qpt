"""
Phase 4: Micro Planning - AVIVA Didactics Service

Handles AVIVA didactic plan generation for training modules.
Uses a hybrid approach:
- Programmatic generation for structure (Start, Duration, Type, AVIVA, Method, Material)
- GenAI generation for content (What column)

Based on Phase4_Micro_Planning_Specification_v1.1.md
"""

import os
import json
from datetime import datetime, time, timedelta
from typing import Dict, Any, List, Optional, Tuple
from openai import OpenAI
from sqlalchemy import text
from flask import current_app
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class Phase4AvivaService:
    """Service for Phase 4 AVIVA Didactics operations"""

    # Level configuration
    LEVEL_NAMES = {
        1: 'Knowing',
        2: 'Understanding',
        4: 'Applying',
        6: 'Mastering'
    }

    # Duration in hours per level
    LEVEL_DURATION_HOURS = {
        1: 1,
        2: 2,
        4: 4,
        6: 8
    }

    # AVIVA phase durations in minutes (defaults and allowed ranges)
    AVIVA_PHASE_DURATIONS = {
        'A_arrive': {'default': 10, 'range': (5, 15)},
        'V_activate': {'default': 15, 'range': (10, 25)},
        'I_inform': {'default': 25, 'range': (20, 35)},
        'V_process': {'default': 30, 'range': (15, 45)},
        'A_evaluate': {'default': 10, 'range': (10, 20)},
        'P_break': {'default': 10, 'range': (10, 15)},
        'P_lunch': {'default': 50, 'range': (45, 50)}
    }

    # AVIVA phase to Type mapping
    AVIVA_TO_TYPE = {
        'A': 'V',  # Arrival -> Lecture (welcome)
        'V': 'U',  # Activate/Process -> Exercise
        'I': 'V',  # Inform -> Lecture
        'P': 'P'   # Pause -> Pause
    }

    # Methods by AVIVA phase and format
    METHODS_BY_PHASE_FORMAT = {
        'A': {
            'default': 'Lecture',
            'seminar': 'Lecture',
            'webinar': 'Presentation',
            'coaching': 'Introduction',
            'wbt': 'Video Introduction'
        },
        'V_activate': {
            'default': 'Discussion',
            'seminar': 'Group Discussion',
            'webinar': 'Poll / Chat Discussion',
            'coaching': 'One-on-One Discussion',
            'wbt': 'Self-Assessment Quiz'
        },
        'I': {
            'default': 'Lecture',
            'seminar': 'Lecture with Q&A',
            'webinar': 'Presentation with Chat',
            'coaching': 'Guided Explanation',
            'wbt': 'Interactive Content'
        },
        'V_process': {
            'default': 'Group Exercise',
            'seminar': 'Group Exercise',
            'webinar': 'Breakout Room Exercise',
            'coaching': 'Guided Practice',
            'wbt': 'Interactive Exercise'
        },
        'A_evaluate': {
            'default': 'Q&A / Feedback',
            'seminar': 'Discussion / Feedback Form',
            'webinar': 'Q&A / Survey',
            'coaching': 'Reflection Discussion',
            'wbt': 'Summary Quiz'
        },
        'P': {
            'default': '-',
            'seminar': '-',
            'webinar': '-',
            'coaching': '-',
            'wbt': '-'
        }
    }

    # Materials by format mode
    MATERIALS_BY_FORMAT = {
        'offline': {
            'A': 'PowerPoint',
            'V_activate': 'Flip-Chart, Moderation Cards',
            'I': 'PowerPoint, Handouts',
            'V_process': 'Exercise Sheets, Case Study',
            'A_evaluate': 'Feedback Form',
            'P': '-'
        },
        'online': {
            'A': 'PowerPoint, Screen Share',
            'V_activate': 'Poll Tool, Chat',
            'I': 'PowerPoint, Screen Share',
            'V_process': 'Digital Collaboration Tool (Miro/Mural)',
            'A_evaluate': 'Survey Tool',
            'P': '-'
        },
        'hybrid': {
            'A': 'PowerPoint',
            'V_activate': 'Whiteboard, Digital Board',
            'I': 'PowerPoint, Shared Documents',
            'V_process': 'Exercise Materials (Physical/Digital)',
            'A_evaluate': 'Feedback Form / Survey',
            'P': '-'
        }
    }

    def __init__(self, db_session):
        """Initialize the service with database session"""
        self.db = db_session
        self.client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        self.model = "gpt-4o-mini"

    def _get_level_name(self, level: int) -> str:
        """Convert level number to display name"""
        return self.LEVEL_NAMES.get(level, f'Level {level}')

    def _minutes_to_time(self, total_minutes: int, start_hour: int = 9) -> str:
        """Convert total minutes from start to HH:MM format"""
        hours = start_hour + (total_minutes // 60)
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"

    def _get_business_day(self, start_date, day_offset: int):
        """Return the Nth business day from start_date, skipping weekends (Sat/Sun).

        Args:
            start_date: datetime.date or datetime.datetime
            day_offset: 0-based offset (0 = start_date itself if it's a weekday)

        Returns:
            datetime.date
        """
        from datetime import date as date_type
        if hasattr(start_date, 'date'):
            current = start_date.date()
        elif isinstance(start_date, date_type):
            current = start_date
        else:
            # Try parsing string
            current = datetime.strptime(str(start_date), '%Y-%m-%d').date()

        # Advance to first weekday if start_date is on a weekend
        while current.weekday() >= 5:  # 5=Sat, 6=Sun
            current += timedelta(days=1)

        # Now advance day_offset business days
        days_added = 0
        while days_added < day_offset:
            current += timedelta(days=1)
            if current.weekday() < 5:
                days_added += 1

        return current

    def _get_format_key(self, format_name: str) -> str:
        """Convert format name to key for mapping lookup"""
        if not format_name:
            return 'default'
        name_lower = format_name.lower()
        if 'seminar' in name_lower:
            return 'seminar'
        elif 'webinar' in name_lower:
            return 'webinar'
        elif 'coaching' in name_lower:
            return 'coaching'
        elif 'wbt' in name_lower or 'web-based' in name_lower:
            return 'wbt'
        elif 'blended' in name_lower:
            return 'seminar'  # Default to seminar for blended
        return 'default'

    def _get_mode_key(self, mode_of_delivery: str) -> str:
        """Convert mode of delivery to key for materials lookup"""
        if not mode_of_delivery:
            return 'offline'
        mode_lower = mode_of_delivery.lower()
        if mode_lower == 'online':
            return 'online'
        elif mode_lower == 'hybrid':
            return 'hybrid'
        return 'offline'

    # =========================================================================
    # AVIVA SEQUENCE GENERATION
    # =========================================================================

    def _generate_aviva_sequence(self, level: int, total_minutes: int) -> List[Dict[str, Any]]:
        """
        Generate AVIVA activity sequence based on level and total duration.

        Returns list of activities with:
        - aviva_phase: A, V, I, or P
        - phase_type: 'arrive', 'activate', 'inform', 'process', 'evaluate', 'break', 'lunch'
        - duration: minutes
        - type: V, U, or P
        """
        activities = []

        # Level 1: 1 hour (60 min) - Compact single-cycle sequence
        # A(5) + V(10) + I(20) + V(15) + A(10) = 60 min
        if level == 1:
            activities = [
                {'aviva_phase': 'A', 'phase_type': 'arrive', 'duration': 5, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'activate', 'duration': 10, 'type': 'U'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 20, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 15, 'type': 'U'},
                {'aviva_phase': 'A', 'phase_type': 'evaluate', 'duration': 10, 'type': 'V'},
            ]

        # Level 2: 2 hours (120 min) - Two I-V cycles with break
        # A(5) + V(10) + I(25) + V(20) + P(10) + I(20) + V(20) + A(10) = 120 min
        elif level == 2:
            activities = [
                {'aviva_phase': 'A', 'phase_type': 'arrive', 'duration': 5, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'activate', 'duration': 10, 'type': 'U'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 25, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 20, 'type': 'U'},
                {'aviva_phase': 'P', 'phase_type': 'break', 'duration': 10, 'type': 'P'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 20, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 20, 'type': 'U'},
                {'aviva_phase': 'A', 'phase_type': 'evaluate', 'duration': 10, 'type': 'V'},
            ]

        # Level 4: 4 hours (240 min) - Half-day workshop, 3 practice rounds, no lunch
        # A(10) + V(15) + I(20) + I(20) + V(30) + P(15) + I(20) + V(35) + P(10) + I(20) + V(30) + A(15) = 240 min
        elif level == 4:
            activities = [
                {'aviva_phase': 'A', 'phase_type': 'arrive', 'duration': 10, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'activate', 'duration': 15, 'type': 'U'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 20, 'type': 'V'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 20, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 30, 'type': 'U'},
                {'aviva_phase': 'P', 'phase_type': 'break', 'duration': 15, 'type': 'P'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 20, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 35, 'type': 'U'},
                {'aviva_phase': 'P', 'phase_type': 'break', 'duration': 10, 'type': 'P'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 20, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 30, 'type': 'U'},
                {'aviva_phase': 'A', 'phase_type': 'evaluate', 'duration': 15, 'type': 'V'},
            ]

        # Level 6: 8 hours (480 min) - Full day with lunch, 4 practice rounds
        # A(15) + V(25) + I(35) + I(30) + V(45) + P(15) + I(30) + V(40) + P(50) + V(15) + I(30) + V(40) + P(15) + I(30) + V(45) + A(20) = 480 min
        elif level == 6:
            activities = [
                {'aviva_phase': 'A', 'phase_type': 'arrive', 'duration': 15, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'activate', 'duration': 25, 'type': 'U'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 35, 'type': 'V'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 30, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 45, 'type': 'U'},
                {'aviva_phase': 'P', 'phase_type': 'break', 'duration': 15, 'type': 'P'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 30, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 40, 'type': 'U'},
                {'aviva_phase': 'P', 'phase_type': 'lunch', 'duration': 50, 'type': 'P'},
                {'aviva_phase': 'V', 'phase_type': 'activate', 'duration': 15, 'type': 'U'},  # Energizer
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 30, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 40, 'type': 'U'},
                {'aviva_phase': 'P', 'phase_type': 'break', 'duration': 15, 'type': 'P'},
                {'aviva_phase': 'I', 'phase_type': 'inform', 'duration': 30, 'type': 'V'},
                {'aviva_phase': 'V', 'phase_type': 'process', 'duration': 45, 'type': 'U'},
                {'aviva_phase': 'A', 'phase_type': 'evaluate', 'duration': 20, 'type': 'V'},
            ]

        else:
            # Default to level 1 pattern
            activities = self._generate_aviva_sequence(1, total_minutes)

        return activities

    def _add_methods_and_materials(
        self,
        activities: List[Dict[str, Any]],
        format_key: str,
        mode_key: str
    ) -> List[Dict[str, Any]]:
        """Add method and material columns based on format and mode"""
        for activity in activities:
            phase = activity['aviva_phase']
            phase_type = activity['phase_type']

            # Determine method lookup key
            if phase == 'V':
                method_key = 'V_activate' if phase_type == 'activate' else 'V_process'
            elif phase == 'A':
                method_key = 'A' if phase_type == 'arrive' else 'A_evaluate'
            else:
                method_key = phase

            # Get method
            method_options = self.METHODS_BY_PHASE_FORMAT.get(method_key, self.METHODS_BY_PHASE_FORMAT.get(phase, {}))
            activity['method'] = method_options.get(format_key, method_options.get('default', 'Lecture'))

            # Get materials - use phase_type for more specific lookup
            if phase == 'V':
                material_key = 'V_activate' if phase_type == 'activate' else 'V_process'
            elif phase == 'A':
                material_key = 'A' if phase_type == 'arrive' else 'A_evaluate'
            else:
                material_key = phase

            materials = self.MATERIALS_BY_FORMAT.get(mode_key, self.MATERIALS_BY_FORMAT['offline'])
            activity['material'] = materials.get(material_key, materials.get(phase, 'PowerPoint'))

        return activities

    def _add_start_times(self, activities: List[Dict[str, Any]], start_hour: int = 9) -> List[Dict[str, Any]]:
        """Add start times to activities"""
        current_minutes = 0
        current_day = 1

        for activity in activities:
            # Check if this is day 2
            if activity.get('day', 1) == 2 and current_day == 1:
                current_day = 2
                current_minutes = 0  # Reset for day 2

            activity['start_time'] = self._minutes_to_time(current_minutes, start_hour)
            current_minutes += activity['duration']

        return activities

    def _add_placeholder_content(
        self,
        activities: List[Dict[str, Any]],
        competency_name: str,
        content_topics: List[str]
    ) -> List[Dict[str, Any]]:
        """Add placeholder content for template export"""
        inform_count = 0
        topic_index = 0

        for activity in activities:
            phase = activity['aviva_phase']
            phase_type = activity['phase_type']

            if phase == 'A' and phase_type == 'arrive':
                activity['content'] = f"[Welcome & Introduction to {competency_name}]"
            elif phase == 'V' and phase_type == 'activate':
                activity['content'] = f"[Prior Knowledge Discussion: {competency_name}]"
            elif phase == 'I':
                if topic_index < len(content_topics):
                    activity['content'] = f"[Content: {content_topics[topic_index]}]"
                    topic_index += 1
                else:
                    activity['content'] = f"[Content Topic {inform_count + 1}]"
                inform_count += 1
            elif phase == 'V' and phase_type == 'process':
                activity['content'] = f"[Practice Exercise {topic_index}]"
            elif phase == 'A' and phase_type == 'evaluate':
                activity['content'] = "[Summary & Evaluation / Q&A]"
            elif phase == 'P':
                if phase_type == 'lunch':
                    activity['content'] = "Lunch Break"
                else:
                    activity['content'] = "Break"

        return activities

    # =========================================================================
    # MODULE DATA RETRIEVAL
    # =========================================================================

    def get_modules_for_aviva(self, organization_id: int) -> Dict[str, Any]:
        """
        Get all training modules ready for AVIVA planning.

        IMPORTANT: Reuses Phase 3's exact data by calling Phase3PlanningService.
        This ensures Phase 4 shows the exact same modules the user configured in Phase 3,
        including the correct view type (competency_level or role_clustered).

        Returns dict with:
        - modules: List of confirmed training modules with AVIVA-specific data
        - view_type: 'competency_level' or 'role_clustered' (from Phase 3 config)
        - scaling_info: Participant scaling information
        """
        # Import Phase 3 service to reuse its logic exactly
        from app.services.phase3_planning_service import Phase3PlanningService
        phase3_service = Phase3PlanningService(self.db)

        # Get Phase 3 config to know which view type was selected
        phase3_config = phase3_service.get_phase3_config(organization_id)
        view_type = phase3_config.get('selected_view', 'competency_level')

        # Get modules using Phase 3's exact logic
        phase3_result = phase3_service.get_training_modules(organization_id, view_type=view_type)

        if phase3_result.get('error'):
            return {
                'modules': [],
                'view_type': view_type,
                'scaling_info': phase3_result.get('scaling_info'),
                'error': phase3_result['error']
            }

        phase3_modules = phase3_result.get('modules', [])
        scaling_info = phase3_result.get('scaling_info', {})

        # Get AVIVA plan status for each module from phase3_training_module
        aviva_status = self._get_aviva_plan_status(organization_id)

        # Filter to only CONFIRMED modules and add Phase 4-specific data
        modules = []
        for m in phase3_modules:
            # Only include confirmed modules with format selection
            if not m.get('confirmed') or not m.get('selected_format_id'):
                continue

            # Build lookup key for AVIVA status
            # Try key with cluster_id first (for role_clustered), then without (for competency_level)
            cluster_id = m.get('cluster_id') or 0
            status_key_with_cluster = f"{m['competency_id']}_{m['target_level']}_{m['pmt_type']}_{cluster_id}"
            status_key_without_cluster = f"{m['competency_id']}_{m['target_level']}_{m['pmt_type']}"

            aviva_info = aviva_status.get(status_key_with_cluster) or aviva_status.get(status_key_without_cluster, {})

            # Get format info
            format_info = self._get_format_info(m.get('selected_format_id'))

            level = m['target_level']
            duration_hours = self.LEVEL_DURATION_HOURS.get(level, 2)
            level_name = self._get_level_name(level)

            # Get module ID - prefer from Phase 3 data, fallback to aviva_status lookup
            module_id = m.get('id') or aviva_info.get('module_id')

            modules.append({
                # Phase 3 data (exactly as displayed in Phase 3)
                'id': module_id,
                'competency_id': m['competency_id'],
                'competency_name': m['competency_name'],
                'target_level': level,
                'level_name': level_name,
                'pmt_type': m['pmt_type'],
                'module_name': m.get('module_name', f"{m['competency_name']} - {level_name}"),
                'users_with_gap': m.get('users_with_gap', 0),
                'estimated_participants': m.get('estimated_participants', 0),
                'roles_needing_training': m.get('roles_needing_training', []),

                # Cluster info (for role_clustered view)
                'cluster_id': m.get('cluster_id'),
                'cluster_name': m.get('cluster_name'),
                'subcluster': m.get('subcluster'),  # 'common' or 'pathway' for Engineers
                'shared_roles_count': m.get('shared_roles_count'),
                'pathway_roles': m.get('pathway_roles'),

                # Format info
                'learning_format_id': m.get('selected_format_id'),
                'learning_format_name': format_info.get('format_name') if format_info else None,
                'mode_of_delivery': format_info.get('mode_of_delivery') if format_info else None,

                # Phase 4-specific data
                'estimated_duration_hours': duration_hours,
                'has_aviva_plan': aviva_info.get('has_aviva_plan', False),
                'learning_objective': m.get('learning_objective', ''),

                # Suitability info from Phase 3
                'suitability': m.get('suitability'),
                'confirmed': True  # Already filtered above
            })

        # Sort by cluster (for role_clustered) then competency, level, pmt_type
        if view_type == 'role_clustered':
            modules.sort(key=lambda m: (
                m.get('cluster_id') or 999,
                m.get('subcluster') or 'z',  # 'common' before 'pathway'
                m['competency_name'],
                m['target_level'],
                m['pmt_type']
            ))
        else:
            modules.sort(key=lambda m: (m['competency_name'], m['target_level'], m['pmt_type']))

        result = {
            'modules': modules,
            'view_type': view_type,
            'scaling_info': scaling_info
        }

        # Pass through Level 1 consolidation metadata from Phase 3
        if phase3_result.get('level1_consolidated'):
            result['level1_consolidated'] = True
            result['level1_modules_removed'] = phase3_result.get('level1_modules_removed', 0)

        return result

    def _get_aviva_plan_status(self, organization_id: int) -> Dict[str, Dict]:
        """Get AVIVA plan status for all modules in the organization.

        Creates multiple key formats to support both views:
        - competency_level view uses: "{comp_id}_{level}_{pmt}"
        - role_clustered view uses: "{comp_id}_{level}_{pmt}_{cluster_id}"
        """
        result = self.db.execute(
            text("""
                SELECT
                    tm.id as module_id,
                    tm.competency_id,
                    tm.target_level,
                    tm.pmt_type,
                    tm.training_program_cluster_id as cluster_id,
                    CASE
                        WHEN EXISTS (SELECT 1 FROM phase4_aviva_plan WHERE training_module_id = tm.id)
                        THEN true ELSE false
                    END as has_aviva_plan
                FROM phase3_training_module tm
                WHERE tm.organization_id = :org_id
                  AND tm.confirmed = true
            """),
            {'org_id': organization_id}
        ).fetchall()

        status_lookup = {}
        for row in result:
            module_info = {
                'module_id': row.module_id,
                'has_aviva_plan': row.has_aviva_plan
            }

            # Key WITH cluster_id (for role_clustered view)
            cluster_id = row.cluster_id or 0
            key_with_cluster = f"{row.competency_id}_{row.target_level}_{row.pmt_type}_{cluster_id}"
            status_lookup[key_with_cluster] = module_info

            # Key WITHOUT cluster_id (for competency_level view)
            # Only add this key if cluster_id is 0/NULL to avoid conflicts
            if not row.cluster_id:
                key_without_cluster = f"{row.competency_id}_{row.target_level}_{row.pmt_type}"
                status_lookup[key_without_cluster] = module_info

        return status_lookup

    def _get_format_info(self, format_id: int) -> Optional[Dict]:
        """Get format information by ID"""
        if not format_id:
            return None

        result = self.db.execute(
            text("""
                SELECT id, format_name, short_name, mode_of_delivery
                FROM learning_format
                WHERE id = :id
            """),
            {'id': format_id}
        ).fetchone()

        if result:
            return {
                'id': result.id,
                'format_name': result.format_name,
                'short_name': result.short_name,
                'mode_of_delivery': result.mode_of_delivery
            }
        return None

    def get_module_preview(self, module_id: int) -> Dict[str, Any]:
        """
        Get detailed preview for a single module including learning objectives and content topics.
        Also retrieves learning objective from Phase 2's generated data.
        For Method/Tool modules, extracts the PMT-specific learning objective.
        """
        # Get module info including cluster data for role-clustered view
        module = self.db.execute(
            text("""
                SELECT
                    tm.id,
                    tm.organization_id,
                    tm.competency_id,
                    c.competency_name,
                    tm.target_level,
                    tm.pmt_type,
                    tm.selected_format_id,
                    lf.format_name,
                    lf.mode_of_delivery,
                    lf.communication_type,
                    tm.estimated_participants,
                    tm.training_program_cluster_id as cluster_id,
                    tpc.training_program_name as cluster_name
                FROM phase3_training_module tm
                JOIN competency c ON tm.competency_id = c.id
                LEFT JOIN learning_format lf ON tm.selected_format_id = lf.id
                LEFT JOIN training_program_cluster tpc ON tm.training_program_cluster_id = tpc.id
                WHERE tm.id = :module_id
            """),
            {'module_id': module_id}
        ).fetchone()

        if not module:
            return None

        level = module.target_level
        pmt_type = module.pmt_type
        duration_hours = self.LEVEL_DURATION_HOURS.get(level, 2)

        # Get learning objective from competency_indicators (template/baseline)
        learning_obj = self.db.execute(
            text("""
                SELECT indicator_en
                FROM competency_indicators
                WHERE competency_id = :comp_id AND level = :level
            """),
            {'comp_id': module.competency_id, 'level': str(level)}
        ).fetchone()

        # Get learning objective from Phase 2 generated data
        # This includes PMT-specific objectives if available
        generated_lo = None
        generated_lo_pmt = None  # PMT-specific objective
        pmt_breakdown = None
        lo_result = self.db.execute(
            text("""
                SELECT objectives_data
                FROM generated_learning_objectives
                WHERE organization_id = :org_id
                ORDER BY generated_at DESC
                LIMIT 1
            """),
            {'org_id': module.organization_id}
        ).fetchone()

        if lo_result and lo_result.objectives_data:
            objectives_data = lo_result.objectives_data
            if isinstance(objectives_data, str):
                objectives_data = json.loads(objectives_data)

            data = objectives_data.get('data', objectives_data)
            main_pyramid = data.get('main_pyramid', {})
            levels = main_pyramid.get('levels', {})
            level_data = levels.get(str(level), {})

            for comp in level_data.get('competencies', []):
                if comp.get('competency_id') == module.competency_id:
                    lo_data = comp.get('learning_objective', {})
                    generated_lo = lo_data.get('objective_text', '')
                    pmt_breakdown = lo_data.get('pmt_breakdown')

                    # Extract PMT-specific learning objective if module has pmt_type
                    if pmt_type in ('method', 'tool') and pmt_breakdown:
                        generated_lo_pmt = pmt_breakdown.get(pmt_type)
                    break

        # Determine the best learning objective to show
        # Priority: PMT-specific > Generated unified > Template
        final_learning_objective = None
        if pmt_type in ('method', 'tool') and generated_lo_pmt:
            final_learning_objective = generated_lo_pmt
        elif generated_lo:
            final_learning_objective = generated_lo
        elif learning_obj:
            final_learning_objective = learning_obj.indicator_en

        # Get content topics from baseline
        content_result = self.db.execute(
            text("""
                SELECT content_topics
                FROM competency_content_baseline
                WHERE competency_id = :comp_id
            """),
            {'comp_id': module.competency_id}
        ).fetchone()

        content_topics = content_result.content_topics if content_result else []

        # Generate suggested AVIVA sequence (preview) with methods
        sequence = self._generate_aviva_sequence(level, duration_hours * 60)

        # Add methods and materials to the sequence
        format_key = self._get_format_key(module.format_name)
        mode_key = self._get_mode_key(module.mode_of_delivery)
        sequence = self._add_methods_and_materials(sequence, format_key, mode_key)

        # Construct module name
        level_name = self._get_level_name(level)
        if pmt_type == 'combined':
            module_name = f"{module.competency_name} - {level_name}"
        else:
            module_name = f"{module.competency_name} - {level_name} - {pmt_type.title()}"

        # Map activity types to human-readable labels
        # V = Vortrag (Frontal/Lecture), U = Ubung (Active/Exercise), P = Pause (Break)
        type_labels = {'V': 'Frontal', 'U': 'Active', 'P': 'Break'}

        return {
            'id': module.id,
            'organization_id': module.organization_id,
            'competency_id': module.competency_id,
            'competency_name': module.competency_name,
            'module_name': module_name,
            'target_level': level,
            'level_name': level_name,
            'pmt_type': pmt_type,
            'learning_format': {
                'id': module.selected_format_id,
                'name': module.format_name,
                'mode_of_delivery': module.mode_of_delivery,
                'communication_type': module.communication_type
            },
            'learning_objective': final_learning_objective,
            'learning_objective_unified': generated_lo,
            'learning_objective_template': learning_obj.indicator_en if learning_obj else None,
            'pmt_breakdown': pmt_breakdown,
            'content_topics': content_topics,
            'estimated_duration_hours': duration_hours,
            'estimated_participants': module.estimated_participants,
            # Cluster info for role-clustered view
            'cluster_id': module.cluster_id,
            'cluster_name': module.cluster_name,
            'subcluster': None,  # Will be populated from Phase 3 modules data if available
            'suggested_activity_count': len(sequence),
            'suggested_sequence': [
                {
                    'aviva': act['aviva_phase'],
                    'phase_type': act['phase_type'],
                    'duration': act['duration'],
                    'type': type_labels.get(act['type'], act['type']),
                    'method': act.get('method', ''),
                    'material': act.get('material', '')
                }
                for act in sequence
            ]
        }

    # =========================================================================
    # AVIVA PLAN GENERATION
    # =========================================================================

    def generate_aviva_plan(
        self,
        module_id: int,
        generation_method: str = 'template',
        start_time_hour: int = 9
    ) -> Dict[str, Any]:
        """
        Generate AVIVA plan for a single module.

        Args:
            module_id: Training module ID
            generation_method: 'template' or 'genai'
            start_time_hour: Starting hour (default 9 for 09:00)

        Returns:
            Complete AVIVA plan with activities
        """
        # Get module preview (includes all needed data)
        preview = self.get_module_preview(module_id)
        if not preview:
            raise ValueError(f"Module {module_id} not found")

        level = preview['target_level']
        total_minutes = preview['estimated_duration_hours'] * 60
        format_key = self._get_format_key(preview['learning_format']['name'])
        mode_key = self._get_mode_key(preview['learning_format']['mode_of_delivery'])

        # Generate AVIVA sequence
        activities = self._generate_aviva_sequence(level, total_minutes)

        # Add methods and materials
        activities = self._add_methods_and_materials(activities, format_key, mode_key)

        # Add start times
        activities = self._add_start_times(activities, start_time_hour)

        # Add content
        if generation_method == 'genai':
            activities = self._generate_genai_content(
                activities,
                preview['competency_name'],
                level,
                preview['learning_objective'],
                preview['content_topics'],
                preview['pmt_type'],
                preview['learning_format']['name']
            )
        else:
            # Template mode - add placeholders
            activities = self._add_placeholder_content(
                activities,
                preview['competency_name'],
                preview['content_topics']
            )

        # Build module name
        module_name = f"{preview['competency_name']} - {preview['level_name']}"
        if preview['pmt_type'] and preview['pmt_type'] != 'combined':
            module_name += f" ({preview['pmt_type'].capitalize()})"

        # Build complete AVIVA plan
        aviva_plan = {
            'module_id': module_id,
            'module_name': module_name,
            'competency_id': preview['competency_id'],
            'competency_name': preview['competency_name'],
            'target_level': level,
            'level_name': preview['level_name'],
            'pmt_type': preview['pmt_type'],
            'learning_format': preview['learning_format']['name'],
            'mode_of_delivery': preview['learning_format']['mode_of_delivery'],
            'total_duration_minutes': total_minutes,
            'learning_objective': preview['learning_objective'],
            'content_topics': preview['content_topics'],
            # Cluster info for role-clustered view
            'cluster_id': preview.get('cluster_id'),
            'cluster_name': preview.get('cluster_name'),
            'subcluster': preview.get('subcluster'),
            'generated_by': generation_method,
            'generated_at': datetime.now().isoformat(),
            'activities': [
                {
                    'row': i + 1,
                    'start_time': act['start_time'],
                    'duration_min': act['duration'],
                    'type': act['type'],
                    'aviva_phase': act['aviva_phase'],
                    'content': act.get('content', ''),
                    'method': act.get('method', ''),
                    'material': act.get('material', ''),
                    'day': act.get('day', 1)
                }
                for i, act in enumerate(activities)
            ]
        }

        return aviva_plan

    def _generate_genai_content(
        self,
        activities: List[Dict[str, Any]],
        competency_name: str,
        level: int,
        learning_objective: str,
        content_topics: List[str],
        pmt_type: str,
        format_name: str
    ) -> List[Dict[str, Any]]:
        """Generate content using OpenAI for each AVIVA activity"""

        # Build sequence description for prompt
        sequence_desc = []
        for i, act in enumerate(activities):
            sequence_desc.append({
                'row': i + 1,
                'aviva_phase': act['aviva_phase'],
                'phase_type': act['phase_type'],
                'duration': act['duration'],
                'type': act['type']
            })

        prompt = f"""You are creating AVIVA didactic content for a Systems Engineering training module.

== MODULE CONTEXT ==
Module: {competency_name} - Level {level} ({self._get_level_name(level)})
PMT Focus: {pmt_type or 'combined'}
Learning Format: {format_name or 'Seminar'}
Total Duration: {sum(a['duration'] for a in activities)} minutes

== LEARNING OBJECTIVE ==
{learning_objective or 'Enable participants to understand and apply this competency.'}

== CONTENT TOPICS (Baseline) ==
{chr(10).join('- ' + t for t in content_topics) if content_topics else '- General competency topics'}

== INSTRUCTIONS ==
Generate specific, actionable content for the "What (Content)" column for each activity row.

Guidelines by AVIVA phase:
- A (Arrive): Welcome, set expectations, preview learning objectives
- V (Activate): Questions or activities to surface existing knowledge
- I (Inform): Specific topic from content list - adapt depth based on level:
  * Level 1: Overview, definitions, awareness
  * Level 2: Explanations, connections, examples
  * Level 4: Practical application, hands-on guidance
  * Level 6: Strategic thinking, teaching others
- V (Process): Concrete exercise description tied to previous content
- A (Evaluate): Summary, key takeaways, reflection
- P (Break/Lunch): Just "Break" or "Lunch Break"

Keep each description to 1-2 sentences. Be specific to {competency_name}.

== ACTIVITY SEQUENCE ==
{json.dumps(sequence_desc, indent=2)}

== OUTPUT FORMAT ==
Return ONLY a valid JSON array with content for each row:
[
  {{"row": 1, "content": "Welcome: Introduction to {competency_name}..."}},
  {{"row": 2, "content": "Discussion: What experience do you have with..."}},
  ...
]
"""

        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are an expert instructional designer for Systems Engineering training programs. Generate clear, specific didactic content."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7,
                max_tokens=2000
            )

            content = response.choices[0].message.content.strip()

            # Parse JSON from response
            # Handle markdown code blocks if present
            if content.startswith('```'):
                content = content.split('```')[1]
                if content.startswith('json'):
                    content = content[4:]
                content = content.strip()

            generated = json.loads(content)

            # Map generated content to activities
            content_map = {item['row']: item['content'] for item in generated}
            for i, act in enumerate(activities):
                row = i + 1
                if row in content_map:
                    act['content'] = content_map[row]
                else:
                    # Fallback to placeholder
                    act['content'] = f"[{act['aviva_phase']} Activity {row}]"

        except Exception as e:
            current_app.logger.error(f"GenAI content generation failed: {e}")
            # Fallback to placeholder content
            activities = self._add_placeholder_content(
                activities,
                competency_name,
                content_topics
            )

        return activities

    # =========================================================================
    # SAVE AND RETRIEVE PLANS
    # =========================================================================

    def save_aviva_plan(self, organization_id: int, module_id: int, aviva_plan: Dict[str, Any]) -> int:
        """Save AVIVA plan to database"""
        result = self.db.execute(
            text("""
                INSERT INTO phase4_aviva_plan (organization_id, training_module_id, generated_by, aviva_content)
                VALUES (:org_id, :module_id, :generated_by, :content)
                ON CONFLICT (training_module_id) DO UPDATE
                SET generated_by = EXCLUDED.generated_by,
                    aviva_content = EXCLUDED.aviva_content,
                    updated_at = CURRENT_TIMESTAMP
                RETURNING id
            """),
            {
                'org_id': organization_id,
                'module_id': module_id,
                'generated_by': aviva_plan['generated_by'],
                'content': json.dumps(aviva_plan)
            }
        )
        self.db.commit()
        return result.fetchone().id

    def get_aviva_plan(self, module_id: int) -> Optional[Dict[str, Any]]:
        """Retrieve saved AVIVA plan for a module"""
        result = self.db.execute(
            text("""
                SELECT aviva_content, generated_by, created_at, updated_at
                FROM phase4_aviva_plan
                WHERE training_module_id = :module_id
            """),
            {'module_id': module_id}
        ).fetchone()

        if result:
            plan = result.aviva_content
            plan['db_generated_by'] = result.generated_by
            plan['created_at'] = result.created_at.isoformat() if result.created_at else None
            plan['updated_at'] = result.updated_at.isoformat() if result.updated_at else None
            return plan
        return None

    # =========================================================================
    # TRAINING DELIVERY SCHEDULE
    # =========================================================================

    DAY_CAPACITY_MINUTES = 480  # 8 hours
    LUNCH_DURATION_MINUTES = 45
    LUNCH_THRESHOLD_MINUTES = 300  # 5 hours -> add lunch
    MERGE_THRESHOLD_MINUTES = 240  # days < 4h eligible for merge

    def _pack_modules_into_days(self, modules: List[Dict], format_label: str) -> List[Dict]:
        """Greedy first-fit decreasing bin-packing of modules into 8h days.

        Args:
            modules: List of module dicts, each must have 'duration_minutes'
            format_label: Label for the format grouping (e.g. 'Seminar Day')

        Returns:
            List of day dicts with 'modules', 'total_minutes', 'format_label'
        """
        if not modules:
            return []

        # Sort largest-first for better packing
        sorted_mods = sorted(modules, key=lambda m: m['duration_minutes'], reverse=True)

        days = []
        for mod in sorted_mods:
            placed = False
            for day in days:
                if day['total_minutes'] + mod['duration_minutes'] <= self.DAY_CAPACITY_MINUTES:
                    day['modules'].append(mod)
                    day['total_minutes'] += mod['duration_minutes']
                    placed = True
                    break
            if not placed:
                days.append({
                    'modules': [mod],
                    'total_minutes': mod['duration_minutes'],
                    'format_label': format_label,
                })

        return days

    def _generate_daily_schedule(
        self,
        aviva_plans: List[Dict[str, Any]],
        view_type: str,
        milestones: List[Dict]
    ) -> Dict[str, Any]:
        """Main orchestrator: groups modules, packs into days, assigns dates.

        Args:
            aviva_plans: Sorted list of AVIVA plan dicts (same as used for export)
            view_type: 'competency_level' or 'role_clustered'
            milestones: Phase 3 timeline milestones list

        Returns:
            Schedule data dict with sections, days, totals
        """
        from itertools import groupby

        # --- Extract rollout dates from milestones ---
        rollout_start = None
        rollout_end = None
        for ms in (milestones or []):
            order = ms.get('order', ms.get('milestone_order'))
            date_str = ms.get('estimated_date', '')
            if order == 4 and date_str:  # Rollout Start
                try:
                    rollout_start = datetime.strptime(date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass
            elif order == 5 and date_str:  # Rollout End
                try:
                    rollout_end = datetime.strptime(date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass

        has_calendar_dates = rollout_start is not None

        # --- Build flat module list with scheduling metadata ---
        def build_schedule_module(plan):
            duration_min = plan.get('total_duration_minutes', 0)
            level = plan.get('target_level', 0)
            return {
                'module_name': plan.get('module_name', ''),
                'competency_name': plan.get('competency_name', ''),
                'target_level': level,
                'level_name': self._get_level_name(level),
                'learning_format': plan.get('learning_format', ''),
                'duration_minutes': duration_min,
                'duration_hours': round(duration_min / 60, 1),
                'cluster_name': plan.get('cluster_name'),
                'subcluster': plan.get('subcluster'),
                'pmt_type': plan.get('pmt_type', ''),
                'estimated_participants': plan.get('estimated_participants', 0),
                'roles_needing_training': plan.get('roles_needing_training', []),
                'has_internal_lunch': level == 6,  # L6 modules are full-day with built-in lunch
            }

        # --- Section grouping ---
        if view_type == 'role_clustered':
            cluster_order = {'SE for Engineers': 0, 'SE for Managers': 1, 'SE for Interfacing Partners': 2}
            sections_raw = {}
            for plan in aviva_plans:
                cname = plan.get('cluster_name', 'Uncategorized')
                if cname not in sections_raw:
                    sections_raw[cname] = []
                sections_raw[cname].append(build_schedule_module(plan))
            # Sort sections by cluster order
            section_list = sorted(sections_raw.items(), key=lambda x: cluster_order.get(x[0], 99))
        else:
            # Single section for competency-level view
            all_mods = [build_schedule_module(p) for p in aviva_plans]
            section_list = [('All Modules', all_mods)]

        # --- Pack each section ---
        schedule_sections = []
        running_day_offset = 0  # For sequential calendar date assignment across packages

        for section_name, section_modules in section_list:
            if not section_modules:
                continue

            # Sub-group by learning format (G2)
            def format_sort_key(m):
                return m.get('learning_format', '') or ''

            sorted_by_format = sorted(section_modules, key=format_sort_key)
            format_groups = {}
            for fmt, grp_iter in groupby(sorted_by_format, key=format_sort_key):
                grp = list(grp_iter)
                # Derive a day label from the format
                fmt_lower = (fmt or '').lower()
                if 'seminar' in fmt_lower:
                    day_label = 'Seminar Day'
                elif 'webinar' in fmt_lower:
                    day_label = 'Webinar Day'
                elif 'coaching' in fmt_lower:
                    day_label = 'Coaching Day'
                elif 'wbt' in fmt_lower or 'web-based' in fmt_lower:
                    day_label = 'WBT Day'
                elif 'blended' in fmt_lower:
                    day_label = 'Blended Day'
                else:
                    day_label = 'Training Day'
                format_groups[fmt] = {'modules': grp, 'day_label': day_label}

            # Pack each format group
            all_days = []
            for fmt, fg in format_groups.items():
                packed = self._pack_modules_into_days(fg['modules'], fg['day_label'])
                all_days.extend(packed)

            # Merge under-filled days (< 4h) across format groups
            under_filled = [d for d in all_days if d['total_minutes'] < self.MERGE_THRESHOLD_MINUTES]
            full_days = [d for d in all_days if d['total_minutes'] >= self.MERGE_THRESHOLD_MINUTES]

            merged_days = []
            while under_filled:
                day = under_filled.pop(0)
                merged = True
                while merged and under_filled:
                    merged = False
                    for i, candidate in enumerate(under_filled):
                        if day['total_minutes'] + candidate['total_minutes'] <= self.DAY_CAPACITY_MINUTES:
                            day['modules'].extend(candidate['modules'])
                            day['total_minutes'] += candidate['total_minutes']
                            day['format_label'] = 'Mixed'
                            under_filled.pop(i)
                            merged = True
                            break
                merged_days.append(day)

            all_days = full_days + merged_days

            # Sort days: full days first (by total_minutes desc), then merged
            all_days.sort(key=lambda d: d['total_minutes'], reverse=True)

            # Assign day numbers and calendar dates, compute per-module times
            section_days = []
            for day_idx, day in enumerate(all_days):
                day_number = day_idx + 1

                # Calendar date
                cal_date = None
                day_of_week = None
                if has_calendar_dates:
                    cal_date = self._get_business_day(rollout_start, running_day_offset + day_idx)
                    day_of_week = cal_date.strftime('%A')

                # Compute per-module start/end times within the day
                current_minutes = 0
                needs_lunch = False
                total_teaching = day['total_minutes']

                # Determine if lunch needed: multi-module days >= 5h without an L6 module
                has_l6 = any(m.get('has_internal_lunch') for m in day['modules'])
                if total_teaching >= self.LUNCH_THRESHOLD_MINUTES and not has_l6 and len(day['modules']) > 1:
                    needs_lunch = True

                day_modules = []
                lunch_inserted = False
                for mod in day['modules']:
                    # Insert lunch if needed around 12:00 (180 min from 9:00)
                    if needs_lunch and not lunch_inserted and current_minutes >= 180:
                        lunch_inserted = True
                        # We don't add lunch to total_minutes of day - it's annotation only

                    start_time = self._minutes_to_time(current_minutes, 9)
                    current_minutes += mod['duration_minutes']
                    end_time = self._minutes_to_time(current_minutes, 9)

                    day_modules.append({
                        'module_name': mod['module_name'],
                        'competency_name': mod.get('competency_name', ''),
                        'start_time': start_time,
                        'end_time': end_time,
                        'duration_hours': mod['duration_hours'],
                        'duration_minutes': mod['duration_minutes'],
                        'target_level': mod['target_level'],
                        'level_name': mod['level_name'],
                        'learning_format': mod['learning_format'],
                        'pmt_type': mod.get('pmt_type', ''),
                        'estimated_participants': mod.get('estimated_participants', 0),
                        'has_internal_lunch': mod.get('has_internal_lunch', False),
                    })

                section_days.append({
                    'day_number': day_number,
                    'calendar_date': cal_date.isoformat() if cal_date else None,
                    'day_of_week': day_of_week,
                    'format_label': day['format_label'],
                    'total_minutes': day['total_minutes'],
                    'total_hours': round(day['total_minutes'] / 60, 1),
                    'needs_lunch': needs_lunch,
                    'modules': day_modules,
                })

            total_section_days = len(section_days)
            total_section_hours = round(sum(d['total_minutes'] for d in section_days) / 60, 1)

            schedule_sections.append({
                'section_name': section_name,
                'total_days': total_section_days,
                'total_hours': total_section_hours,
                'days': section_days,
            })

            # Advance running offset for next package (G3 sequential scheduling)
            running_day_offset += total_section_days

        grand_total_days = sum(s['total_days'] for s in schedule_sections)
        grand_total_hours = round(sum(s['total_hours'] for s in schedule_sections), 1)

        # Determine schedule period display
        schedule_period = None
        if has_calendar_dates and grand_total_days > 0:
            first_date = schedule_sections[0]['days'][0]['calendar_date'] if schedule_sections and schedule_sections[0]['days'] else None
            last_section = schedule_sections[-1] if schedule_sections else None
            last_date = last_section['days'][-1]['calendar_date'] if last_section and last_section['days'] else None
            if first_date and last_date:
                schedule_period = f"{first_date} to {last_date}"

        return {
            'schedule_sections': schedule_sections,
            'grand_total_days': grand_total_days,
            'grand_total_hours': grand_total_hours,
            'has_calendar_dates': has_calendar_dates,
            'schedule_period': schedule_period,
            'rollout_start': rollout_start.isoformat() if rollout_start else None,
            'rollout_end': rollout_end.isoformat() if rollout_end else None,
        }

    def _write_schedule_sheet(self, ws, schedule_data: Dict[str, Any], view_type: str, milestones=None):
        """Write the Training Delivery Schedule Excel sheet.

        Args:
            ws: openpyxl worksheet
            schedule_data: Output from _generate_daily_schedule()
            view_type: 'competency_level' or 'role_clustered'
            milestones: Optional list of timeline milestone dicts
        """
        from openpyxl.utils import get_column_letter

        # Styles
        TITLE_FONT = Font(size=16, bold=True)
        SUBTITLE_FONT = Font(size=11, bold=True)
        LABEL_FONT = Font(bold=True)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        DAY_HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        DAY_HEADER_FONT = Font(bold=True, size=10)
        LUNCH_FONT = Font(italic=True, color="888888")
        TOTAL_FONT = Font(bold=True, italic=True)

        PKG_FILLS = {
            'SE for Engineers': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
            'SE for Managers': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'SE for Interfacing Partners': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        }
        PKG_HEADER_FONT = Font(bold=True, size=11)

        # Column layout
        col_headers = ['Day #', 'Date', 'Time', 'Module', 'Level', 'Format', 'Duration']
        col_widths = [7, 16, 14, 45, 14, 28, 11]
        num_cols = len(col_headers)

        # Set column widths
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # ===== TITLE =====
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        ws['A1'] = 'Training Delivery Schedule'
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='center')
        row = 3

        # ===== SUMMARY =====
        grand_days = schedule_data.get('grand_total_days', 0)
        grand_hours = schedule_data.get('grand_total_hours', 0)
        has_dates = schedule_data.get('has_calendar_dates', False)
        period = schedule_data.get('schedule_period')

        ws[f'A{row}'] = 'Total Training Days:'
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'B{row}'] = grand_days
        row += 1

        ws[f'A{row}'] = 'Total Training Hours:'
        ws[f'A{row}'].font = LABEL_FONT
        hours_display = f"{grand_hours}h"
        if grand_hours >= 8:
            hours_display += f" (~{grand_hours / 8:.1f} training days)"
        ws[f'B{row}'] = hours_display
        ws.merge_cells(f'B{row}:D{row}')
        row += 1

        ws[f'A{row}'] = 'Schedule Period:'
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'B{row}'] = period if period else 'Dates not yet determined'
        ws.merge_cells(f'B{row}:D{row}')
        row += 2

        # ===== PER-SECTION BLOCKS =====
        for section in schedule_data.get('schedule_sections', []):
            section_name = section['section_name']
            section_days = section['total_days']
            section_hours = section['total_hours']

            # Section header
            pkg_fill = PKG_FILLS.get(section_name, PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'))
            ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
            cell = ws[f'A{row}']
            cell.value = f"{section_name} - Training Schedule"
            cell.font = PKG_HEADER_FONT
            cell.fill = pkg_fill
            cell.alignment = Alignment(vertical='center')
            row += 1

            # Section summary
            ws[f'A{row}'] = f"{section_days} training day{'s' if section_days != 1 else ''} | {section_hours}h total"
            ws[f'A{row}'].font = Font(italic=True, size=9, color='606266')
            ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
            row += 1

            # Column headers
            for col, header in enumerate(col_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

            # Days
            for day in section.get('days', []):
                day_num = day['day_number']
                cal_date = day.get('calendar_date', '')
                day_of_week = day.get('day_of_week', '')
                fmt_label = day.get('format_label', 'Training Day')
                total_h = day.get('total_hours', 0)

                # Day header row
                if cal_date and day_of_week:
                    day_header = f"Day {day_num} -- {day_of_week[:3]}, {cal_date} | {fmt_label} | {total_h}h"
                else:
                    day_header = f"Day {day_num} | {fmt_label} | {total_h}h"

                ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
                cell = ws[f'A{row}']
                cell.value = day_header
                cell.fill = DAY_HEADER_FILL
                cell.font = DAY_HEADER_FONT
                row += 1

                # Module rows
                needs_lunch = day.get('needs_lunch', False)
                lunch_inserted = False

                for mod in day.get('modules', []):
                    start_t = mod.get('start_time', '')
                    end_t = mod.get('end_time', '')

                    # Insert lunch annotation before this module if needed
                    if needs_lunch and not lunch_inserted:
                        # Parse start minutes to check if we're past 12:00 (180 min)
                        try:
                            parts = start_t.split(':')
                            start_min = (int(parts[0]) - 9) * 60 + int(parts[1])
                        except (ValueError, IndexError):
                            start_min = 0
                        if start_min >= 180:
                            lunch_inserted = True
                            ws.cell(row=row, column=1, value='').border = thin_border
                            date_display = cal_date if cal_date else ''
                            ws.cell(row=row, column=2, value=date_display).border = thin_border
                            ws.cell(row=row, column=3, value='12:00 - 12:45').border = thin_border
                            lunch_cell = ws.cell(row=row, column=4, value='Lunch Break')
                            lunch_cell.border = thin_border
                            lunch_cell.font = LUNCH_FONT
                            ws.cell(row=row, column=5, value='').border = thin_border
                            ws.cell(row=row, column=6, value='').border = thin_border
                            ws.cell(row=row, column=7, value='45 min').border = thin_border
                            for c in range(1, num_cols + 1):
                                ws.cell(row=row, column=c).alignment = Alignment(horizontal='center')
                            ws.cell(row=row, column=4).alignment = Alignment(horizontal='left')
                            row += 1

                    # Module data row
                    date_display = cal_date if cal_date else ''
                    level_display = f"L{mod['target_level']} {mod['level_name']}"

                    ws.cell(row=row, column=1, value=f"Day {day_num}").border = thin_border
                    ws.cell(row=row, column=2, value=date_display).border = thin_border
                    ws.cell(row=row, column=3, value=f"{start_t} - {end_t}").border = thin_border
                    ws.cell(row=row, column=4, value=mod['module_name']).border = thin_border
                    ws.cell(row=row, column=5, value=level_display).border = thin_border
                    ws.cell(row=row, column=6, value=mod.get('learning_format', '')).border = thin_border
                    ws.cell(row=row, column=7, value=f"{mod['duration_hours']}h").border = thin_border

                    # Alignment
                    for c in [1, 2, 3, 5, 7]:
                        ws.cell(row=row, column=c).alignment = Alignment(horizontal='center')

                    row += 1

                # Insert lunch at end if still needed and not yet inserted
                if needs_lunch and not lunch_inserted:
                    lunch_inserted = True
                    ws.cell(row=row, column=1, value='').border = thin_border
                    ws.cell(row=row, column=2, value=cal_date if cal_date else '').border = thin_border
                    ws.cell(row=row, column=3, value='12:00 - 12:45').border = thin_border
                    lunch_cell = ws.cell(row=row, column=4, value='Lunch Break')
                    lunch_cell.border = thin_border
                    lunch_cell.font = LUNCH_FONT
                    ws.cell(row=row, column=5, value='').border = thin_border
                    ws.cell(row=row, column=6, value='').border = thin_border
                    ws.cell(row=row, column=7, value='45 min').border = thin_border
                    for c in range(1, num_cols + 1):
                        ws.cell(row=row, column=c).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=4).alignment = Alignment(horizontal='left')
                    row += 1

            # Section total
            ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
            ws[f'A{row}'] = f"Total: {section_days} day{'s' if section_days != 1 else ''}, {section_hours}h"
            ws[f'A{row}'].font = TOTAL_FONT
            ws[f'A{row}'].alignment = Alignment(horizontal='right')
            row += 2

        # ===== IMPLEMENTATION TIMELINE =====
        if milestones:
            ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
            ws[f'A{row}'] = 'Implementation Timeline'
            ws[f'A{row}'].font = SUBTITLE_FONT
            row += 1

            ms_headers = ['#', 'Milestone', 'Target Date', 'Quarter', 'Description']
            ms_col_count = len(ms_headers)
            ms_col_widths = [5, 35, 15, 12, 45]
            for col, hdr in enumerate(ms_headers, 1):
                cell = ws.cell(row=row, column=col, value=hdr)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Only widen columns if needed (schedule columns may already be wider)
                cur_width = ws.column_dimensions[get_column_letter(col)].width or 0
                if ms_col_widths[col - 1] > cur_width:
                    ws.column_dimensions[get_column_letter(col)].width = ms_col_widths[col - 1]
            row += 1

            for idx, milestone in enumerate(milestones, 1):
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=milestone.get('milestone_name', milestone.get('name', ''))).border = thin_border
                ws.cell(row=row, column=3, value=milestone.get('estimated_date', '')).border = thin_border
                ws.cell(row=row, column=4, value=milestone.get('quarter', '')).border = thin_border
                desc_cell = ws.cell(row=row, column=5, value=milestone.get('milestone_description', milestone.get('description', '')))
                desc_cell.border = thin_border
                desc_cell.alignment = Alignment(wrap_text=True)

                for col in [1, 3, 4]:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                row += 1

    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================

    def export_aviva_to_excel(self, aviva_plans: List[Dict[str, Any]], organization_id: int = None) -> io.BytesIO:
        """
        Export AVIVA plans to Excel file.

        Args:
            aviva_plans: List of AVIVA plan dictionaries
            organization_id: Optional organization ID to include Phase 3 summary data

        Returns:
            BytesIO buffer containing Excel file
        """
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Styles
        TITLE_FONT = Font(size=16, bold=True)
        SUBTITLE_FONT = Font(size=11, bold=True)
        LABEL_FONT = Font(bold=True)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")  # Blue Grey
        module_header_fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
        aviva_header_fill = PatternFill(start_color="78909C", end_color="78909C", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Level names mapping
        LEVEL_NAMES = {1: 'Knowing', 2: 'Understanding', 4: 'Applying', 6: 'Mastering'}

        row = 1

        # ===== TITLE =====
        ws.merge_cells('A1:H1')
        ws['A1'] = 'SE-QPT Phase 4: AVIVA Didactic Plans Export'
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='center')
        row = 3

        # ===== PHASE 3 SUMMARY DATA (if organization_id provided) =====
        org_name = f'Organization_{organization_id}' if organization_id else 'Unknown'
        view_type = 'competency_level'
        all_strategies = []
        target_group_size = 0
        format_distribution = {}
        milestones = []

        maturity_score = 0
        maturity_level = 1
        all_strategy_objs = []
        target_group_range = ''

        if organization_id:
            # Get organization name and maturity
            org_result = self.db.execute(
                text("SELECT organization_name, maturity_score FROM organization WHERE id = :org_id"),
                {'org_id': organization_id}
            ).fetchone()
            if org_result:
                org_name = org_result.organization_name
                maturity_score = org_result.maturity_score or 0
                if maturity_score >= 80: maturity_level = 5
                elif maturity_score >= 60: maturity_level = 4
                elif maturity_score >= 40: maturity_level = 3
                elif maturity_score >= 20: maturity_level = 2
                else: maturity_level = 1

            # Get Phase 3 output data
            from app.services.phase3_planning_service import Phase3PlanningService
            phase3_service = Phase3PlanningService(self.db)
            try:
                output = phase3_service.get_phase3_output(organization_id)
                config = output.get('config', {})
                summary = output.get('summary', {})
                timeline = output.get('timeline', {})

                view_type = config.get('selected_view', 'competency_level')
                all_strategies = summary.get('all_strategies', [summary.get('strategy_name', '')])
                target_group_size = summary.get('target_group_size', 0)
                target_group_range = summary.get('target_group_range', '')
                format_distribution = summary.get('format_distribution', {})
                milestones = timeline.get('milestones', [])
                all_strategy_objs = phase3_service._get_all_strategies(organization_id)
            except Exception as e:
                logging.warning(f"Could not get Phase 3 data: {e}")

        # Style for description/annotation rows
        DESC_FONT = Font(italic=True, size=9, color='606266')
        PHASE_FONT = Font(italic=True, size=9, color='909399')

        # Helper: write an overview item with optional description and phase origin
        def write_overview_item(r, label, value, phase_origin=None, description=None):
            ws[f'A{r}'] = label
            ws[f'A{r}'].font = LABEL_FONT
            ws[f'B{r}'] = value
            ws.merge_cells(f'B{r}:F{r}')
            if phase_origin:
                ws[f'G{r}'] = phase_origin
                ws[f'G{r}'].font = PHASE_FONT
            r += 1
            if description:
                ws[f'B{r}'] = description
                ws[f'B{r}'].font = DESC_FONT
                ws.merge_cells(f'B{r}:H{r}')
                r += 1
            return r

        # ===== OVERVIEW SECTION =====
        ws[f'A{row}'] = 'Overview'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        # Organization
        row = write_overview_item(row, 'Organization:', org_name,
            description='The organization for which this SE qualification plan was created.')

        # SE Maturity Level
        row = write_overview_item(row, 'SE Maturity Level:', f'Level {maturity_level}/5 (Score: {maturity_score:.1f}/100)',
            phase_origin='(Phase 1, Task 3)',
            description='Organization SE maturity assessed via the Maturity Assessment questionnaire.')

        # Target Group Size
        target_display = f"{target_group_size} participants"
        if target_group_range and target_group_range != 'Unknown':
            target_display += f" ({target_group_range})"
        row = write_overview_item(row, 'Target Group Size:', target_display,
            phase_origin='(Phase 1, Task 1)',
            description='Number of employees in the target group to be qualified in SE competencies.')

        # Selected Strategies (with descriptions)
        from app.strategy_selection_engine import SE_TRAINING_STRATEGIES
        strategy_desc_map = {s['name']: s.get('description', '') for s in SE_TRAINING_STRATEGIES}
        strategy_detail_map = {s['name']: s for s in SE_TRAINING_STRATEGIES}

        strategy_names_str = ', '.join(all_strategies) if all_strategies else 'Not Selected'
        if all_strategy_objs:
            labeled = []
            for s in all_strategy_objs:
                lbl = f"{s['name']} (Primary)" if s.get('is_primary') else s['name']
                labeled.append(lbl)
            strategy_names_str = ', '.join(labeled)
        row = write_overview_item(row, 'Selected Strategies:', strategy_names_str,
            phase_origin='(Phase 1, Task 2)',
            description='Qualification strategies selected to guide the training program design.')

        # Individual strategy descriptions
        for s_name in all_strategies:
            desc = strategy_desc_map.get(s_name, '')
            detail = strategy_detail_map.get(s_name, {})
            if desc:
                ws[f'B{row}'] = f"{s_name}:"
                ws[f'B{row}'].font = Font(bold=True, size=9)
                row += 1
                ws[f'B{row}'] = desc
                ws[f'B{row}'].font = DESC_FONT
                ws.merge_cells(f'B{row}:H{row}')
                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                row += 1
                details_parts = []
                if detail.get('targetAudience'):
                    details_parts.append(f"Target: {detail['targetAudience']}")
                if detail.get('qualificationLevel'):
                    details_parts.append(f"Level: {detail['qualificationLevel']}")
                if detail.get('duration'):
                    details_parts.append(f"Duration: {detail['duration']}")
                if details_parts:
                    ws[f'B{row}'] = ' | '.join(details_parts)
                    ws[f'B{row}'].font = PHASE_FONT
                    ws.merge_cells(f'B{row}:H{row}')
                    row += 1

        # Training View
        is_role_clustered = view_type == 'role_clustered'
        view_label = 'Role-Clustered' if is_role_clustered else 'Competency-Level'
        if is_role_clustered:
            view_desc = 'Modules grouped into 3 training packages (SE for Engineers, SE for Managers, SE for Interfacing Partners) based on role cluster assignments.'
        else:
            view_desc = 'Modules grouped by the 16 SE competency areas, each containing sub-modules at identified competency levels.'
        row = write_overview_item(row, 'Training View:', view_label,
            phase_origin='(Phase 3, Task 1)',
            description=view_desc)

        # Total Training Modules
        row = write_overview_item(row, 'Total Training Modules:', len(aviva_plans),
            phase_origin='(Phase 3, Task 2)',
            description='Training modules identified from competency gap analysis and learning format assignment.')

        # Total Estimated Duration
        total_duration = sum(plan.get('total_duration_minutes', 0) for plan in aviva_plans) / 60
        duration_display = f"{total_duration:.1f} hours"
        if total_duration >= 8:
            duration_display += f" (~{total_duration / 8:.1f} training days)"
        row = write_overview_item(row, 'Total Est. Duration:', duration_display,
            phase_origin='(Derived)',
            description='Cumulative estimated training duration across all selected modules.')

        # Format Distribution
        if format_distribution:
            format_str = ', '.join([f"{k}: {v}" for k, v in format_distribution.items()])
            row = write_overview_item(row, 'Format Distribution:', format_str,
                phase_origin='(Phase 3, Task 2)',
                description='Distribution of learning formats selected for training delivery.')

        # Export Date
        ws[f'A{row}'] = 'Export Date:'
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        row += 2

        # ===== GENAI DISCLAIMER =====
        DISCLAIMER_FONT = Font(size=9, italic=True, color="666666")
        disclaimer_text = (
            "Disclaimer: This document contains content generated using Generative AI (GenAI) "
            "and may contain inaccuracies or errors. It is intended as a reference document and "
            "starting point. Review and modifications may be necessary to ensure accuracy and "
            "alignment with your specific organizational requirements."
        )
        ws.merge_cells(f'A{row}:H{row}')
        disclaimer_cell = ws[f'A{row}']
        disclaimer_cell.value = disclaimer_text
        disclaimer_cell.font = DISCLAIMER_FONT
        disclaimer_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 40
        row += 2

        # Set column widths for overview area
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14

        # ===== MODULES TABLE =====
        ws[f'A{row}'] = 'Training Modules Overview'
        ws[f'A{row}'].font = SUBTITLE_FONT
        row += 1

        # Check if role-clustered view
        is_role_clustered = view_type == 'role_clustered'

        # Sort plans - by cluster for role_clustered, by competency_id otherwise
        if is_role_clustered:
            # Define cluster order
            cluster_order = {'SE for Engineers': 0, 'SE for Managers': 1, 'SE for Interfacing Partners': 2}
            sorted_plans = sorted(aviva_plans, key=lambda p: (
                cluster_order.get(p.get('cluster_name', ''), 99),
                p.get('subcluster') or 'z',  # 'common' before 'pathway'
                p.get('competency_id', 0)
            ))
        else:
            sorted_plans = sorted(aviva_plans, key=lambda p: p.get('competency_id', 0))

        # Module summary headers - different for role-clustered vs competency-level
        if is_role_clustered:
            summary_headers = ['#', 'Training Program', 'Module Type', 'Module', 'Level', 'Format', 'Roles', 'Est. Participants', 'Duration (h)', 'Activities']
            col_widths = [4, 20, 13, 40, 14, 32, 45, 12, 10, 9]
            center_cols = [1, 5, 8, 9, 10]
        else:
            summary_headers = ['#', 'Module', 'Level', 'Format', 'Roles', 'Est. Participants', 'Duration (h)', 'Activities']
            col_widths = [4, 42, 14, 30, 50, 12, 10, 9]
            center_cols = [1, 3, 6, 7, 8]

        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            _cl = get_column_letter(col)
            ws.column_dimensions[_cl].width = max(col_widths[col - 1], ws.column_dimensions[_cl].width or 0)
        row += 1

        # Style constants for section headers
        COMP_HEADER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        COMP_HEADER_FONT = Font(bold=True, italic=True, size=10)
        PKG_ENGINEERS_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        PKG_MANAGERS_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        PKG_PARTNERS_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        PKG_HEADER_FONT = Font(bold=True, size=11)
        SUBCLUSTER_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        SUBCLUSTER_FONT = Font(bold=True, italic=True, size=10, color='606060')

        pkg_fills = {
            'SE for Engineers': PKG_ENGINEERS_FILL,
            'SE for Managers': PKG_MANAGERS_FILL,
            'SE for Interfacing Partners': PKG_PARTNERS_FILL,
        }

        num_cols = len(summary_headers)

        def write_section_header_aviva(row_num, label, fill, font):
            merge_range = f'A{row_num}:{get_column_letter(num_cols)}{row_num}'
            ws.merge_cells(merge_range)
            cell = ws[f'A{row_num}']
            cell.value = label
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical='center')
            return row_num + 1

        def format_module_display(plan):
            competency_name = plan.get('competency_name', '')
            pmt_type = plan.get('pmt_type', '')
            if pmt_type and pmt_type.lower() not in ['combined', '', 'null', 'none']:
                return f"{competency_name} ({pmt_type.capitalize()})"
            return competency_name

        def write_plan_row_rc(row_num, plan, idx):
            cluster_name = plan.get('cluster_name', 'Uncategorized')
            subcluster = plan.get('subcluster')
            if 'Engineer' in cluster_name:
                module_type = 'Common Base' if subcluster == 'common' else 'Role-Specific'
            else:
                module_type = '-'
            level_name = LEVEL_NAMES.get(plan.get('target_level', 0), f"Level {plan.get('target_level', 0)}")
            roles = plan.get('roles_needing_training', [])
            roles_str = ', '.join(roles) if roles and isinstance(roles, list) else ''
            ws.cell(row=row_num, column=1, value=idx).border = thin_border
            ws.cell(row=row_num, column=2, value=cluster_name).border = thin_border
            ws.cell(row=row_num, column=3, value=module_type).border = thin_border
            ws.cell(row=row_num, column=4, value=format_module_display(plan)).border = thin_border
            ws.cell(row=row_num, column=5, value=level_name).border = thin_border
            ws.cell(row=row_num, column=6, value=plan.get('learning_format', '')).border = thin_border
            roles_cell = ws.cell(row=row_num, column=7, value=roles_str)
            roles_cell.border = thin_border
            roles_cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=row_num, column=8, value=plan.get('estimated_participants', 0)).border = thin_border
            ws.cell(row=row_num, column=9, value=round(plan['total_duration_minutes'] / 60, 1)).border = thin_border
            ws.cell(row=row_num, column=10, value=len(plan.get('activities', []))).border = thin_border
            for col in center_cols:
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
            return row_num + 1

        def write_plan_row_cl(row_num, plan, idx):
            level_name = LEVEL_NAMES.get(plan.get('target_level', 0), f"Level {plan.get('target_level', 0)}")
            roles = plan.get('roles_needing_training', [])
            roles_str = ', '.join(roles) if roles and isinstance(roles, list) else ''
            ws.cell(row=row_num, column=1, value=idx).border = thin_border
            ws.cell(row=row_num, column=2, value=format_module_display(plan)).border = thin_border
            ws.cell(row=row_num, column=3, value=level_name).border = thin_border
            ws.cell(row=row_num, column=4, value=plan.get('learning_format', '')).border = thin_border
            roles_cell = ws.cell(row=row_num, column=5, value=roles_str)
            roles_cell.border = thin_border
            roles_cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=row_num, column=6, value=plan.get('estimated_participants', 0)).border = thin_border
            ws.cell(row=row_num, column=7, value=round(plan['total_duration_minutes'] / 60, 1)).border = thin_border
            ws.cell(row=row_num, column=8, value=len(plan.get('activities', []))).border = thin_border
            for col in center_cols:
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
            return row_num + 1

        from itertools import groupby

        def write_comp_groups_aviva(row_num, section_plans, idx_counter, write_fn):
            sorted_by_comp = sorted(section_plans, key=lambda p: p.get('competency_id', 0))
            for _, grp_iter in groupby(sorted_by_comp, key=lambda p: p.get('competency_id', 0)):
                grp = list(grp_iter)
                comp_name = grp[0].get('competency_name', 'Unknown')
                count = len(grp)
                label = f"    {comp_name} ({count} sub-module{'s' if count != 1 else ''})"
                row_num = write_section_header_aviva(row_num, label, COMP_HEADER_FILL, COMP_HEADER_FONT)
                for plan in grp:
                    idx_counter[0] += 1
                    row_num = write_fn(row_num, plan, idx_counter[0])
            return row_num

        if is_role_clustered:
            idx_counter = [0]
            # Group sorted_plans by cluster_name
            for cluster_name, cluster_iter in groupby(sorted_plans, key=lambda p: p.get('cluster_name', 'Uncategorized')):
                cluster_plans = list(cluster_iter)
                pkg_fill = pkg_fills.get(cluster_name, PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'))
                row = write_section_header_aviva(row, cluster_name, pkg_fill, PKG_HEADER_FONT)

                if 'Engineer' in cluster_name:
                    common_plans = [p for p in cluster_plans if p.get('subcluster') == 'common']
                    pathway_plans = [p for p in cluster_plans if p.get('subcluster') != 'common']

                    if common_plans:
                        cb_label = f"  Common Base ({len(common_plans)} module{'s' if len(common_plans) != 1 else ''})"
                        row = write_section_header_aviva(row, cb_label, SUBCLUSTER_FILL, SUBCLUSTER_FONT)
                        row = write_comp_groups_aviva(row, common_plans, idx_counter, write_plan_row_rc)

                    if pathway_plans:
                        rs_label = f"  Role-Specific Pathways ({len(pathway_plans)} module{'s' if len(pathway_plans) != 1 else ''})"
                        row = write_section_header_aviva(row, rs_label, SUBCLUSTER_FILL, SUBCLUSTER_FONT)
                        row = write_comp_groups_aviva(row, pathway_plans, idx_counter, write_plan_row_rc)
                else:
                    row = write_comp_groups_aviva(row, cluster_plans, idx_counter, write_plan_row_rc)
        else:
            idx_counter = [0]
            sorted_by_comp = sorted(sorted_plans, key=lambda p: p.get('competency_id', 0))
            for _, grp_iter in groupby(sorted_by_comp, key=lambda p: p.get('competency_id', 0)):
                grp = list(grp_iter)
                comp_name = grp[0].get('competency_name', 'Unknown')
                count = len(grp)
                label = f"{comp_name} ({count} sub-module{'s' if count != 1 else ''})"
                row = write_section_header_aviva(row, label, COMP_HEADER_FILL, COMP_HEADER_FONT)
                for plan in grp:
                    idx_counter[0] += 1
                    row = write_plan_row_cl(row, plan, idx_counter[0])

        row += 2

        # ===== TRAINING DELIVERY SCHEDULE SHEET =====
        schedule_data = self._generate_daily_schedule(sorted_plans, view_type, milestones)
        ws_schedule = wb.create_sheet(title="Training Schedule", index=1)
        self._write_schedule_sheet(ws_schedule, schedule_data, view_type, milestones=milestones)

        # ===== THIRD SHEET: ALL AVIVA PLANS COMBINED =====
        ws_aviva = wb.create_sheet(title="AVIVA Plans")
        aviva_row = 1

        # Sheet title
        ws_aviva.merge_cells('A1:G1')
        ws_aviva['A1'] = 'Detailed AVIVA Didactic Plans'
        ws_aviva['A1'].font = TITLE_FONT
        ws_aviva['A1'].alignment = Alignment(horizontal='center')
        aviva_row = 3

        # GenAI disclaimer on AVIVA sheet
        ws_aviva.merge_cells(f'A{aviva_row}:G{aviva_row}')
        aviva_disclaimer_cell = ws_aviva[f'A{aviva_row}']
        aviva_disclaimer_cell.value = disclaimer_text
        aviva_disclaimer_cell.font = DISCLAIMER_FONT
        aviva_disclaimer_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws_aviva.row_dimensions[aviva_row].height = 40
        aviva_row += 2

        # AVIVA headers
        aviva_headers = ['Start', 'Min', 'Type', 'AVIVA', 'What (Content)', 'How (Method)', 'Material']
        aviva_col_widths = [18, 8, 10, 15, 50, 25, 45]  # First column wider for labels, Material wider

        # Set column widths for AVIVA sheet
        for col, width in enumerate(aviva_col_widths, 1):
            ws_aviva.column_dimensions[get_column_letter(col)].width = width

        # Write each module's AVIVA plan sequentially
        for plan in sorted_plans:
            # Module header section
            level_num = plan.get('target_level', 0)
            level_name = LEVEL_NAMES.get(level_num, f'Level {level_num}')

            ws_aviva.merge_cells(f'A{aviva_row}:G{aviva_row}')
            cell = ws_aviva[f'A{aviva_row}']
            cell.value = f"Module: {plan['module_name']}"
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = module_header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            aviva_row += 1

            # Module details - add Training Program for role-clustered view
            if is_role_clustered and plan.get('cluster_name'):
                ws_aviva[f'A{aviva_row}'] = 'Training Program:'
                ws_aviva[f'A{aviva_row}'].font = LABEL_FONT
                ws_aviva[f'B{aviva_row}'] = plan.get('cluster_name', '')
                ws_aviva[f'D{aviva_row}'] = 'Module Type:'
                ws_aviva[f'D{aviva_row}'].font = LABEL_FONT
                cluster_name = plan.get('cluster_name', '')
                subcluster = plan.get('subcluster')
                if 'Engineer' in cluster_name:
                    module_type = 'Common Base' if subcluster == 'common' else 'Role-Specific'
                else:
                    module_type = '-'
                ws_aviva[f'E{aviva_row}'] = module_type
                aviva_row += 1

            ws_aviva[f'A{aviva_row}'] = 'Level:'
            ws_aviva[f'A{aviva_row}'].font = LABEL_FONT
            ws_aviva[f'B{aviva_row}'] = level_name
            ws_aviva[f'C{aviva_row}'] = 'Format:'
            ws_aviva[f'C{aviva_row}'].font = LABEL_FONT
            ws_aviva[f'D{aviva_row}'] = plan.get('learning_format', 'N/A')
            ws_aviva[f'E{aviva_row}'] = 'Duration:'
            ws_aviva[f'E{aviva_row}'].font = LABEL_FONT
            ws_aviva[f'F{aviva_row}'] = f"{plan['total_duration_minutes']} min ({round(plan['total_duration_minutes'] / 60, 1)} h)"
            aviva_row += 1

            ws_aviva[f'A{aviva_row}'] = 'Learning Objective:'
            ws_aviva[f'A{aviva_row}'].font = LABEL_FONT
            ws_aviva.merge_cells(f'B{aviva_row}:G{aviva_row}')
            lo_cell = ws_aviva[f'B{aviva_row}']
            lo_cell.value = plan.get('learning_objective', 'N/A')
            lo_cell.alignment = Alignment(wrap_text=True)
            aviva_row += 1

            # Content Topics if available
            content_topics = plan.get('content_topics', [])
            if content_topics:
                ws_aviva[f'A{aviva_row}'] = 'Content Topics:'
                ws_aviva[f'A{aviva_row}'].font = LABEL_FONT
                ws_aviva.merge_cells(f'B{aviva_row}:G{aviva_row}')
                topics_cell = ws_aviva[f'B{aviva_row}']
                topics_cell.value = ', '.join(content_topics)
                topics_cell.alignment = Alignment(wrap_text=True)
                aviva_row += 1

            aviva_row += 1  # Empty row before table

            # AVIVA table headers
            for col, header in enumerate(aviva_headers, 1):
                cell = ws_aviva.cell(row=aviva_row, column=col, value=header)
                cell.font = header_font
                cell.fill = aviva_header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            aviva_row += 1

            # AVIVA activities
            activities = plan.get('activities', [])
            for activity in activities:
                ws_aviva.cell(row=aviva_row, column=1, value=activity.get('start_time', '')).border = thin_border
                ws_aviva.cell(row=aviva_row, column=2, value=activity.get('duration_min', 0)).border = thin_border
                ws_aviva.cell(row=aviva_row, column=3, value=activity.get('type', '')).border = thin_border
                ws_aviva.cell(row=aviva_row, column=4, value=activity.get('aviva_phase', '')).border = thin_border

                content_cell = ws_aviva.cell(row=aviva_row, column=5, value=activity.get('content', ''))
                content_cell.border = thin_border
                content_cell.alignment = Alignment(wrap_text=True)

                ws_aviva.cell(row=aviva_row, column=6, value=activity.get('method', '')).border = thin_border
                ws_aviva.cell(row=aviva_row, column=7, value=activity.get('material', '')).border = thin_border

                # Center align certain columns
                for col in [1, 2, 3, 4]:
                    ws_aviva.cell(row=aviva_row, column=col).alignment = Alignment(horizontal='center')

                aviva_row += 1

            # Add 3 empty rows between modules
            aviva_row += 3

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # =========================================================================
    # PHASE 4 CONFIG
    # =========================================================================

    def get_phase4_config(self, organization_id: int) -> Dict[str, Any]:
        """Get or create Phase 4 configuration"""
        result = self.db.execute(
            text("""
                SELECT id, task1_status, task2_status, aviva_generation_method,
                       created_at, updated_at
                FROM phase4_config
                WHERE organization_id = :org_id
            """),
            {'org_id': organization_id}
        ).fetchone()

        if result:
            return {
                'id': result.id,
                'task1_status': result.task1_status,
                'task2_status': result.task2_status,
                'aviva_generation_method': result.aviva_generation_method,
                'created_at': result.created_at.isoformat() if result.created_at else None,
                'updated_at': result.updated_at.isoformat() if result.updated_at else None
            }

        # Create default config
        self.db.execute(
            text("""
                INSERT INTO phase4_config (organization_id)
                VALUES (:org_id)
                ON CONFLICT (organization_id) DO NOTHING
            """),
            {'org_id': organization_id}
        )
        self.db.commit()

        return {
            'id': None,
            'task1_status': 'not_started',
            'task2_status': 'not_started',
            'aviva_generation_method': None,
            'created_at': None,
            'updated_at': None
        }

    def update_phase4_config(self, organization_id: int, updates: Dict[str, Any]) -> Dict[str, Any]:
        """Update Phase 4 configuration"""
        set_clauses = []
        params = {'org_id': organization_id}

        for key in ['task1_status', 'task2_status', 'aviva_generation_method']:
            if key in updates:
                set_clauses.append(f"{key} = :{key}")
                params[key] = updates[key]

        if set_clauses:
            query = f"""
                UPDATE phase4_config
                SET {', '.join(set_clauses)}, updated_at = CURRENT_TIMESTAMP
                WHERE organization_id = :org_id
            """
            self.db.execute(text(query), params)
            self.db.commit()

        return self.get_phase4_config(organization_id)
